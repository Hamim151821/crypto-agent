from openai import OpenAI
import requests
from dotenv import load_dotenv
import os
import pandas as pd
import ta
import openpyxl
from datetime import datetime
from plyer import notification
import schedule
import time
import threading
import gspread
from google.oauth2.service_account import Credentials
import json
import re

# Load API key dari file .env
load_dotenv()
client = OpenAI(
    api_key=os.getenv("OPENROUTER_API_KEY"),
    base_url="https://openrouter.ai/api/v1"
)
CRYPTOPANIC_TOKEN = os.getenv("CRYPTOPANIC_TOKEN")
NEWS_API_KEY = os.getenv("NEWS_API_KEY")
MARKETAUX_API_KEY = os.getenv("MARKETAUX_API_KEY")

# ==============================
# DEFAULT WEIGHTS & CONFIG
# ==============================
DEFAULT_WEIGHTS = {
    "rsi": 1.0,
    "macd": 1.0,
    "trend": 1.0,
    "volume": 1.0,
    "sentimen": 1.0
}
DEFAULT_MODAL = 1_000_000  # Rp 1.000.000

# ==============================
# FUNGSI AMBIL DATA CRYPTO
# ==============================
def get_crypto_price(nama_crypto):
    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {
        "ids": nama_crypto,
        "vs_currencies": "idr,usd",
        "include_24hr_change": "true"
    }
    try:
        response = requests.get(url, params=params, timeout=10)
        data = response.json()
        
        # Validasi response
        if not data or nama_crypto not in data:
            return None
        
        # Validasi data lengkap
        if not all(k in data[nama_crypto] for k in ['idr', 'usd']):
            return None
            
        return data
    except Exception as e:
        print(f"Error get_crypto_price: {e}")
        return None

# ==============================
# FUNGSI AMBIL DATA SAHAM
# ==============================
def get_stock_price(kode_saham):
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{kode_saham}"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        response = requests.get(url, headers=headers, timeout=10)
        data = response.json()
        
        # Validasi response structure
        if "chart" not in data or "result" not in data["chart"]:
            return None
            
        result = data["chart"]["result"]
        if not result:
            return None
            
        meta = result[0].get("meta", {})
        if not meta or "regularMarketPrice" not in meta:
            return None
            
        price = meta["regularMarketPrice"]
        currency = meta.get("currency", "USD")
        
        if not price or price <= 0:
            return None
            
        return {"harga": price, "mata_uang": currency, "kode": kode_saham}
    except Exception as e:
        print(f"Error get_stock_price: {e}")
        return None

# ==============================
# FOREIGN FLOW DATA (Placeholder)
# ==============================
def get_foreign_flow(kode_saham):
    """
    Ambil data Net Foreign Buy/Sell (placeholder - perlu integrasi API nyata)
    Untuk saat ini, return berdasarkan price action sebagai proxy
    """
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{kode_saham}"
        params = {"interval": "1d", "range": "5d"}
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, params=params, headers=headers, timeout=10)
        data = response.json()
        
        if "chart" not in data or "result" not in data["chart"]:
            return None
        
        result = data["chart"]["result"]
        if not result:
            return None
        
        quotes = result[0].get("indicators", {}).get("quote", [{}])[0]
        closes = quotes.get("close", [])
        
        if len(closes) < 2:
            return None
        
        # Proxy: Harga naik = kemungkinan net foreign buy, harga turun = net sell
        # Ini hanya pendekatan, untuk data sebenarnya perlu API khusus
        recent_change = (closes[-1] - closes[0]) / closes[0] * 100 if closes[0] > 0 else 0
        
        if recent_change > 1:
            net_flow = "Net Buy"
            status = "Accumulation"
        elif recent_change < -1:
            net_flow = "Net Sell"
            status = "Distribution"
        else:
            net_flow = "Net Neutral"
            status = "Waiting"
        
        return {
            "net_flow": net_flow,
            "status": status,
            "change_5d": round(recent_change, 2),
            "available": True
        }
    except Exception as e:
        print(f"Error get_foreign_flow: {e}")
        return {"net_flow": "N/A", "status": "Data Tidak Tersedia", "available": False}

# ==============================
# FUNGSI AMBIL BERITA CRYPTO
# ==============================
def get_crypto_news(nama_crypto):
    url = "https://newsapi.org/v2/everything"
    params = {
        "q": nama_crypto,
        "apiKey": NEWS_API_KEY,
        "language": "en",
        "sortBy": "publishedAt",
        "pageSize": 5
    }
    try:
        response = requests.get(url, params=params)
        data = response.json()
        berita = []
        for item in data.get("articles", [])[:5]:
            berita.append({
                "judul": item["title"],
                "sumber": item["source"]["name"]
            })
        return berita
    except:
        return []
    
# ==============================
# FUNGSI BERITA MARKETAUX
# ==============================
def get_marketaux_news(simbol):
    url = "https://api.marketaux.com/v1/news/all"
    params = {
        "symbols": simbol,
        "api_token": MARKETAUX_API_KEY,
        "limit": 3,
        "language": "en"
    }
    try:
        response = requests.get(url, params=params)
        data = response.json()
        berita = []
        for item in data.get("data", [])[:3]:
            berita.append({
                "judul": item["title"],
                "sumber": item.get("source", "Unknown")
            })
        return berita
    except:
        return []

# ==============================
# FUNGSI AMBIL BERITA SAHAM
# ==============================
def get_stock_news(kode_saham):
    nama_map = {
    "BBRI.JK": "Bank Rakyat Indonesia BRI saham",
    "BBCA.JK": "Bank Central Asia",
    "TLKM.JK": "Telkom Indonesia saham",
    "GOTO.JK": "GoTo Gojek Tokopedia saham",
    "BBNI.JK": "Bank Negara Indonesia BNI saham",
    "BMRI.JK": "Bank Mandiri saham",
    "ASII.JK": "Astra International saham",
    "UNVR.JK": "Unilever Indonesia saham",
    "AAPL": "Apple stock",
    "TSLA": "Tesla stock",
    "NVDA": "Nvidia stock",
    "GOOGL": "Google Alphabet stock",
    "MSFT": "Microsoft stock",
    "AMZN": "Amazon stock"
}
    
    keyword = nama_map.get(kode_saham, kode_saham)
    
    # Ambil dari NewsAPI
    url = "https://newsapi.org/v2/everything"
    params = {
        "q": f'"{keyword}"',
        "apiKey": NEWS_API_KEY,
        "language": "en",
        "sortBy": "publishedAt",
        "pageSize": 3
    }
    berita = []
    try:
        response = requests.get(url, params=params)
        data = response.json()
        for item in data.get("articles", [])[:3]:
            berita.append({
                "judul": item["title"],
                "sumber": item["source"]["name"]
            })
    except:
        pass
    
    # Untuk saham Indonesia, skip Marketaux karena tidak relevan
    if not kode_saham.endswith(".JK"):
        marketaux_berita = get_marketaux_news(kode_saham)
        berita.extend(marketaux_berita)
    
    return berita[:5]  # Maksimal 5 berita

# ==============================
# HELPER: MARKET CONDITION
# ==============================
def detect_market_condition(df, ma50, ma200):
    """
    Tentukan kondisi market dengan LOGIKA BARU:
    1. Jika ADX < 20 → SIDEWAYS/CHOPPY (PRIORITAS PERTAMA)
    2. Jika harga < MA20 DAN < MA50 DAN < MA200 → BEARISH (strict)
    3. Jika harga > MA20 DAN > MA50 DAN > MA200 → BULLISH (strict)
    4. Lainnya → berdasarkan MA50 vs MA200
    
    Priority: ADX > Price Position > MA Crossover
    """
    if len(df) < 20:
        return "UNKNOWN"
    
    close = df["close"]
    current_price = close.iloc[-1]
    
    # Ambil MA values
    ma20 = df["ma20"].iloc[-1] if "ma20" in df.columns and pd.notna(df["ma20"].iloc[-1]) else None
    
    # Hitung ADX jika tersedia
    adx = None
    if "adx" in df.columns:
        adx = df["adx"].iloc[-1]
    
    # === ADX CHECK PRIORITAS PERTAMA ===
    if adx is not None and adx < 20:
        return "SIDEWAYS (CHOPPY)"
        
    # === LOGIKA BARU: PRICE POSITION CHECK ===
    if current_price > 0 and ma20 is not None and ma50 is not None and ma200 is not None:
        if current_price < ma20 and current_price < ma50 and current_price < ma200:
            return "TRENDING DOWN"  
        if current_price > ma20 and current_price > ma50 and current_price > ma200:
            return "TRENDING UP"  
    
    # === FALLBACK: MA CROSSOVER ===
    # Hitung volatilitas dari daily returns (std dev 20 hari terakhir)
    returns = close.pct_change().dropna()
    if len(returns) >= 20:
        volatility = returns.iloc[-20:].std() * 100
    else:
        volatility = returns.std() * 100
    
    is_volatile = volatility > 3.0
    
    # Tentukan trend dari MA50 vs MA200
    if ma50 is not None and ma200 is not None and current_price > 0:
        ma_diff_pct = abs(ma50 - ma200) / current_price * 100
        
        if ma_diff_pct < 1.5:
            trend = "SIDEWAYS"
        elif ma50 > ma200:
            trend = "TRENDING UP"
        else:
            trend = "TRENDING DOWN"
    else:
        # Fallback: gunakan price range 20 hari
        recent = close.iloc[-20:]
        if current_price > 0:
            price_range_pct = (recent.max() - recent.min()) / current_price * 100
            if price_range_pct < 5:
                trend = "SIDEWAYS"
            elif ma50 and ma200 and ma50 > ma200:
                trend = "TRENDING UP"
            elif ma50 and ma200:
                trend = "TRENDING DOWN"
            else:
                trend = "UNKNOWN"
        else:
            trend = "UNKNOWN"
    
    # Tambahkan VOLATILE jika volatilitas tinggi
    if is_volatile:
        return f"{trend} + VOLATILE"
    
    return trend

# ==============================
# INDIKATOR TEKNIKAL CRYPTO (ENHANCED)
# ==============================
def get_crypto_indicators(nama_crypto):
    try:
        url = f"https://api.coingecko.com/api/v3/coins/{nama_crypto}/market_chart"
        # Ambil 300 hari untuk MA200
        params = {"vs_currency": "usd", "days": "300", "interval": "daily"}
        response = requests.get(url, params=params, timeout=30)
        data = response.json()
        
        # Validasi response
        if not data or "prices" not in data or not data["prices"]:
            return None
            
        prices = [x[1] for x in data.get("prices", [])]
        volumes = [x[1] for x in data.get("total_volumes", [])]
        
        if not prices or len(prices) < 50:
            return None
        
        # Get high/low for ATR calculation if available
        highs = None
        lows = None
        if "market_caps" not in data:
            # If no highs/lows, estimate from close
            highs = prices
            lows = prices
        else:
            # Can't get high/low from coinGecko, use close as approximation
            highs = prices
            lows = prices
        
        df = pd.DataFrame({
            "close": prices,
            "volume": volumes[:len(prices)] if volumes else [0] * len(prices),
            "high": highs[:len(prices)] if highs else prices,
            "low": lows[:len(prices)] if lows else prices
        })
        
        # === BASIC INDICATORS ===
        
        # RSI (14)
        df["rsi"] = ta.momentum.RSIIndicator(df["close"], window=14).rsi()
        
        # MACD
        macd_ind = ta.trend.MACD(df["close"])
        df["macd"] = macd_ind.macd()
        df["macd_signal"] = macd_ind.macd_signal()
        df["macd_histogram"] = df["macd"] - df["macd_signal"]
        
        # MA50 & MA200
        df["ma50"] = df["close"].rolling(window=50).mean()
        df["ma200"] = df["close"].rolling(window=200).mean()
        df["ma20"] = df["close"].rolling(window=20).mean()
        
        # Volume moving average (20 hari)
        df["vol_avg"] = df["volume"].rolling(window=20).mean()
        
        # === ADVANCED INDICATORS ===
        
        # Bollinger Bands (20, 2)
        bollinger = ta.volatility.BollingerBands(df["close"], window=20, window_dev=2)
        df["bb_upper"] = bollinger.bollinger_hband()
        df["bb_middle"] = bollinger.bollinger_mavg()
        df["bb_lower"] = bollinger.bollinger_lband()
        
        # ATR (14) - Average True Range
        df["atr"] = ta.volatility.AverageTrueRange(df["high"], df["low"], df["close"], window=14).average_true_range()
        
        # Stochastic Oscillator (14, 3, 3)
        stoch = ta.momentum.StochasticOscillator(df["high"], df["low"], df["close"], window=14, smooth_window=3)
        df["stoch_k"] = stoch.stoch()
        df["stoch_d"] = stoch.stoch_signal()
        
        # ADX - Average Directional Index
        df["adx"] = ta.trend.ADXIndicator(df["high"], df["low"], df["close"], window=14).adx()
        
        # OBV - On Balance Volume
        df["obv"] = ta.volume.OnBalanceVolumeIndicator(df["close"], df["volume"]).on_balance_volume()
        
        # VWAP (approximation for daily)
        df["vwap"] = (df["close"] * df["volume"]).rolling(window=20).sum() / df["volume"].rolling(window=20).sum()
        
        # Ambil nilai terakhir
        rsi = round(df["rsi"].iloc[-1], 2) if pd.notna(df["rsi"].iloc[-1]) else 50.0
        macd_val = round(df["macd"].iloc[-1], 4) if pd.notna(df["macd"].iloc[-1]) else 0
        macd_sig = round(df["macd_signal"].iloc[-1], 4) if pd.notna(df["macd_signal"].iloc[-1]) else 0
        macd_hist = round(df["macd_histogram"].iloc[-1], 4) if pd.notna(df["macd_histogram"].iloc[-1]) else 0
        ma50 = round(df["ma50"].iloc[-1], 4) if pd.notna(df["ma50"].iloc[-1]) else None
        ma200 = round(df["ma200"].iloc[-1], 4) if pd.notna(df["ma200"].iloc[-1]) else None
        ma20 = round(df["ma20"].iloc[-1], 4) if pd.notna(df["ma20"].iloc[-1]) else None
        
        # OBV values
        obv = df["obv"].iloc[-1] if pd.notna(df["obv"].iloc[-1]) else 0
        obv_prev = df["obv"].iloc[-2] if len(df) > 1 and pd.notna(df["obv"].iloc[-2]) else 0
        
        # OBV Divergence Detection
        # Bullish Divergence: Price makes Lower Lows but OBV makes Higher Lows = Accumulation
        # Bearish Divergence: Price makes Higher Highs but OBV makes Lower Highs = Distribution
        obv_divergence = "NONE"
        
        if len(df) >= 10:
            # Get last 10 closes and OBV
            recent_closes_10 = df["close"].iloc[-10:].tolist()
            recent_obv_10 = df["obv"].iloc[-10:].tolist()
            
            # Check price direction (last 5 vs previous 5)
            price_first_5_avg = sum(recent_closes_10[:5]) / 5
            price_last_5_avg = sum(recent_closes_10[5:]) / 5
            obv_first_5_avg = sum(recent_obv_10[:5]) / 5
            obv_last_5_avg = sum(recent_obv_10[5:]) / 5
            
            price_falling = price_last_5_avg < price_first_5_avg
            price_rising = price_last_5_avg > price_first_5_avg
            obv_rising = obv_last_5_avg > obv_first_5_avg
            obv_falling = obv_last_5_avg < obv_first_5_avg
            
            # Bullish Divergence: Price down but OBV up (Accumulation)
            if price_falling and obv_rising:
                obv_divergence = "BULLISH (Accumulation)"
            # Bearish Divergence: Price up but OBV down (Distribution)
            elif price_rising and obv_falling:
                obv_divergence = "BEARISH (Distribution)"
        
        # Bollinger Bands
        bb_upper = round(df["bb_upper"].iloc[-1], 4) if pd.notna(df["bb_upper"].iloc[-1]) else None
        bb_middle = round(df["bb_lower"].iloc[-1], 4) if pd.notna(df["bb_middle"].iloc[-1]) else None
        bb_lower = round(df["bb_lower"].iloc[-1], 4) if pd.notna(df["bb_lower"].iloc[-1]) else None
        
        # ATR
        atr = round(df["atr"].iloc[-1], 4) if pd.notna(df["atr"].iloc[-1]) else 0
        
        # Stochastic
        stoch_k = round(df["stoch_k"].iloc[-1], 2) if pd.notna(df["stoch_k"].iloc[-1]) else 50.0
        stoch_d = round(df["stoch_d"].iloc[-1], 2) if pd.notna(df["stoch_d"].iloc[-1]) else 50.0
        
        # ADX
        adx = round(df["adx"].iloc[-1], 2) if pd.notna(df["adx"].iloc[-1]) else 0
        
        # VWAP
        vwap = round(df["vwap"].iloc[-1], 4) if pd.notna(df["vwap"].iloc[-1]) else None
        
        current_vol = df["volume"].iloc[-1] if pd.notna(df["volume"].iloc[-1]) else 0
        avg_vol = df["vol_avg"].iloc[-1] if pd.notna(df["vol_avg"].iloc[-1]) else current_vol
        
        # Support & Resistance (dari 20 hari terakhir)
        recent_closes = df["close"].iloc[-20:]
        support = round(recent_closes.min(), 4)
        resistance = round(recent_closes.max(), 4)
        current_price = round(df["close"].iloc[-1], 4)
        
        # Fibonacci Retracement levels
        fib_high = max(recent_closes)
        fib_low = min(recent_closes)
        fib_range = fib_high - fib_low
        fib_382 = round(fib_low + fib_range * 0.382, 4)
        fib_500 = round(fib_low + fib_range * 0.500, 4)
        fib_618 = round(fib_low + fib_range * 0.618, 4)
        
        # Pivot Points (Classic)
        pivot = round((fib_high + fib_low + current_price) / 3, 4)
        r1 = round(2 * pivot - fib_low, 4)
        s1 = round(2 * pivot - fib_high, 4)
        r2 = round(pivot + fib_range, 4)
        s2 = round(pivot - fib_range, 4)
        
        # === INTERPRETASI ===
        
        # RSI
        if rsi < 30:
            rsi_status = "OVERSOLD"
        elif rsi > 70:
            rsi_status = "OVERBOUGHT"
        else:
            rsi_status = "NORMAL"
        
        # MACD
        macd_status = "BULLISH" if macd_val > macd_sig else "BEARISH"
        
        # MACD Histogram direction
        macd_hist_status = "MENGUAT" if macd_hist > 0 else "MELEMAH"
        
        # Trend (MA50 vs MA200)
        trend_status = "NEUTRAL"
        if ma50 is not None and ma200 is not None:
            if ma50 > ma200:
                trend_status = "BULLISH"
            else:
                trend_status = "BEARISH"
        
        # Volume
        vol_ratio = round(current_vol / avg_vol, 2) if avg_vol > 0 else 1.0
        if vol_ratio > 1.2:
            volume_status = "TINGGI"
        elif vol_ratio < 0.8:
            volume_status = "RENDAH"
        else:
            volume_status = "NORMAL"
        
        # Bollinger Band position
        if current_price < bb_lower:
            bb_position = "DI BAWAH LOWER BAND"
        elif current_price > bb_upper:
            bb_position = "DI ATAS UPPER BAND"
        elif current_price > bb_middle:
            bb_position = "UPPER HALF"
        else:
            bb_position = "LOWER HALF"
        
        # Stochastic
        if stoch_k < 20:
            stoch_status = "OVERSOLD"
        elif stoch_k > 80:
            stoch_status = "OVERBOUGHT"
        else:
            stoch_status = "NORMAL"
        
        # ADX - Trend strength
        if adx > 25:
            adx_status = "TREND KUAT"
        elif adx > 15:
            adx_status = "TREND MODERATE"
        else:
            adx_status = "TREND LEMAH"
        
        # Market condition
        market_condition = detect_market_condition(df, ma50, ma200)
        
        # Data quality score
        data_quality = min(100, (len(prices) / 300) * 100) if len(prices) < 300 else 100
        
        return {
            "rsi": rsi,
            "rsi_status": rsi_status,
            "macd": macd_val,
            "macd_signal": macd_sig,
            "macd_histogram": macd_hist,
            "macd_status": macd_status,
            "macd_hist_status": macd_hist_status,
            "ma50": ma50,
            "ma200": ma200,
            "ma20": ma20,
            "trend_status": trend_status,
            "volume_current": round(current_vol, 2),
            "volume_avg": round(avg_vol, 2),
            "volume_ratio": vol_ratio,
            "volume_status": volume_status,
            "support": support,
            "resistance": resistance,
            "current_price": current_price,
            "market_condition": market_condition,
            # Advanced indicators
            "bb_upper": bb_upper,
            "bb_middle": bb_middle,
            "bb_lower": bb_lower,
            "bb_position": bb_position,
            "atr": atr,
            "stoch_k": stoch_k,
            "stoch_d": stoch_d,
            "stoch_status": stoch_status,
            "adx": adx,
            "adx_status": adx_status,
            "vwap": vwap,
            # Fibonacci & Pivot
            "fib_382": fib_382,
            "fib_500": fib_500,
            "fib_618": fib_618,
            "pivot": pivot,
            "r1": r1,
            "s1": s1,
            "r2": r2,
            "s2": s2,
            # Volume Flow - OBV
            "obv": round(obv, 2) if obv is not None else 0,
            "obv_prev": round(obv_prev, 2) if obv_prev is not None else 0,
            "obv_divergence": obv_divergence,
            # Data quality
            "data_quality": data_quality
        }
    except Exception as e:
        print(f"⚠️ Error crypto indicators: {e}")
        return None

# ==============================
# INDIKATOR TEKNIKAL SAHAM (ENHANCED)
# ==============================
def get_stock_indicators(kode_saham):
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{kode_saham}"
        params = {"interval": "1d", "range": "1y"}
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, params=params, headers=headers, timeout=30)
        data = response.json()
        
        # Validasi response structure
        if "chart" not in data or "result" not in data["chart"]:
            return None
            
        result = data["chart"]["result"]
        if not result:
            return None
            
        result = result[0]
        quotes = result.get("indicators", {}).get("quote", [{}])[0]
        
        closes = quotes.get("close", [])
        volumes = quotes.get("volume", [])
        highs = quotes.get("high", [])
        lows = quotes.get("low", [])
        
        # Filter None values — simpan index valid
        valid_data = []
        for i in range(len(closes)):
            if closes[i] is not None:
                valid_data.append({
                    "close": closes[i],
                    "volume": volumes[i] if i < len(volumes) and volumes[i] is not None else 0,
                    "high": highs[i] if i < len(highs) and highs[i] is not None else closes[i],
                    "low": lows[i] if i < len(lows) and lows[i] is not None else closes[i]
                })
        
        if not valid_data or len(valid_data) < 50:
            return None
        
        df = pd.DataFrame(valid_data)
        
        # === BASIC INDICATORS ===
        
        # RSI (14)
        df["rsi"] = ta.momentum.RSIIndicator(df["close"], window=14).rsi()
        
        # MACD
        macd_ind = ta.trend.MACD(df["close"])
        df["macd"] = macd_ind.macd()
        df["macd_signal"] = macd_ind.macd_signal()
        df["macd_histogram"] = df["macd"] - df["macd_signal"]
        
        # MA50 & MA200
        df["ma50"] = df["close"].rolling(window=50).mean()
        df["ma200"] = df["close"].rolling(window=200).mean()
        df["ma20"] = df["close"].rolling(window=20).mean()
        
        # Volume moving average (20 hari)
        df["vol_avg"] = df["volume"].rolling(window=20).mean()
        
        # === ADVANCED INDICATORS ===
        
        # Bollinger Bands (20, 2)
        bollinger = ta.volatility.BollingerBands(df["close"], window=20, window_dev=2)
        df["bb_upper"] = bollinger.bollinger_hband()
        df["bb_middle"] = bollinger.bollinger_mavg()
        df["bb_lower"] = bollinger.bollinger_lband()
        
        # ATR (14) - Average True Range
        df["atr"] = ta.volatility.AverageTrueRange(df["high"], df["low"], df["close"], window=14).average_true_range()
        
        # Stochastic Oscillator (14, 3, 3)
        stoch = ta.momentum.StochasticOscillator(df["high"], df["low"], df["close"], window=14, smooth_window=3)
        df["stoch_k"] = stoch.stoch()
        df["stoch_d"] = stoch.stoch_signal()
        
        # ADX - Average Directional Index
        df["adx"] = ta.trend.ADXIndicator(df["high"], df["low"], df["close"], window=14).adx()
        
        # OBV - On Balance Volume
        df["obv"] = ta.volume.OnBalanceVolumeIndicator(df["close"], df["volume"]).on_balance_volume()
        
        # VWAP
        df["vwap"] = (df["close"] * df["volume"]).rolling(window=20).sum() / df["volume"].rolling(window=20).sum()
        
        # Ambil nilai terakhir
        rsi = round(df["rsi"].iloc[-1], 2) if pd.notna(df["rsi"].iloc[-1]) else 50.0
        macd_val = round(df["macd"].iloc[-1], 4) if pd.notna(df["macd"].iloc[-1]) else 0
        macd_sig = round(df["macd_signal"].iloc[-1], 4) if pd.notna(df["macd_signal"].iloc[-1]) else 0
        macd_hist = round(df["macd_histogram"].iloc[-1], 4) if pd.notna(df["macd_histogram"].iloc[-1]) else 0
        ma50 = round(df["ma50"].iloc[-1], 4) if pd.notna(df["ma50"].iloc[-1]) else None
        ma200 = round(df["ma200"].iloc[-1], 4) if pd.notna(df["ma200"].iloc[-1]) else None
        ma20 = round(df["ma20"].iloc[-1], 4) if pd.notna(df["ma20"].iloc[-1]) else None
        
        # Bollinger Bands
        bb_upper = round(df["bb_upper"].iloc[-1], 4) if pd.notna(df["bb_upper"].iloc[-1]) else None
        bb_middle = round(df["bb_middle"].iloc[-1], 4) if pd.notna(df["bb_middle"].iloc[-1]) else None
        bb_lower = round(df["bb_lower"].iloc[-1], 4) if pd.notna(df["bb_lower"].iloc[-1]) else None
        
        # ATR
        atr = round(df["atr"].iloc[-1], 4) if pd.notna(df["atr"].iloc[-1]) else 0
        
        # Stochastic
        stoch_k = round(df["stoch_k"].iloc[-1], 2) if pd.notna(df["stoch_k"].iloc[-1]) else 50.0
        stoch_d = round(df["stoch_d"].iloc[-1], 2) if pd.notna(df["stoch_d"].iloc[-1]) else 50.0
        
        # ADX
        adx = round(df["adx"].iloc[-1], 2) if pd.notna(df["adx"].iloc[-1]) else 0
        
        # VWAP
        vwap = round(df["vwap"].iloc[-1], 4) if pd.notna(df["vwap"].iloc[-1]) else None
        
        # OBV values and divergence
        obv = df["obv"].iloc[-1] if pd.notna(df["obv"].iloc[-1]) else 0
        obv_prev = df["obv"].iloc[-2] if len(df) > 1 and pd.notna(df["obv"].iloc[-2]) else 0
        
        # OBV Divergence Detection
        obv_divergence = "NONE"
        
        if len(df) >= 10:
            recent_closes_10 = df["close"].iloc[-10:].tolist()
            recent_obv_10 = df["obv"].iloc[-10:].tolist()
            
            price_first_5_avg = sum(recent_closes_10[:5]) / 5
            price_last_5_avg = sum(recent_closes_10[5:]) / 5
            obv_first_5_avg = sum(recent_obv_10[:5]) / 5
            obv_last_5_avg = sum(recent_obv_10[5:]) / 5
            
            price_falling = price_last_5_avg < price_first_5_avg
            price_rising = price_last_5_avg > price_first_5_avg
            obv_rising = obv_last_5_avg > obv_first_5_avg
            obv_falling = obv_last_5_avg < obv_first_5_avg
            
            if price_falling and obv_rising:
                obv_divergence = "BULLISH (Accumulation)"
            elif price_rising and obv_falling:
                obv_divergence = "BEARISH (Distribution)"
        
        current_vol = df["volume"].iloc[-1] if pd.notna(df["volume"].iloc[-1]) else 0
        avg_vol = df["vol_avg"].iloc[-1] if pd.notna(df["vol_avg"].iloc[-1]) else current_vol
        
        # Support & Resistance (dari high/low 20 hari terakhir)
        recent_high = df["high"].iloc[-20:]
        recent_low = df["low"].iloc[-20:]
        support = round(recent_low.min(), 4)
        resistance = round(recent_high.max(), 4)
        current_price = round(df["close"].iloc[-1], 4)
        
        # Fibonacci Retracement levels
        fib_high = recent_high.max()
        fib_low = recent_low.min()
        fib_range = fib_high - fib_low
        fib_382 = round(fib_low + fib_range * 0.382, 4)
        fib_500 = round(fib_low + fib_range * 0.500, 4)
        fib_618 = round(fib_low + fib_range * 0.618, 4)
        
        # Pivot Points (Classic)
        pivot = round((fib_high + fib_low + current_price) / 3, 4)
        r1 = round(2 * pivot - fib_low, 4)
        s1 = round(2 * pivot - fib_high, 4)
        r2 = round(pivot + fib_range, 4)
        s2 = round(pivot - fib_range, 4)
        
        # === INTERPRETASI ===
        
        # RSI
        if rsi < 30:
            rsi_status = "OVERSOLD"
        elif rsi > 70:
            rsi_status = "OVERBOUGHT"
        else:
            rsi_status = "NORMAL"
        
        # MACD
        macd_status = "BULLISH" if macd_val > macd_sig else "BEARISH"
        
        # MACD Histogram direction
        macd_hist_status = "MENGUAT" if macd_hist > 0 else "MELEMAH"
        
        # Trend
        trend_status = "NEUTRAL"
        if ma50 is not None and ma200 is not None:
            if ma50 > ma200:
                trend_status = "BULLISH"
            else:
                trend_status = "BEARISH"
        
        # Volume
        vol_ratio = round(current_vol / avg_vol, 2) if avg_vol > 0 else 1.0
        if vol_ratio > 1.2:
            volume_status = "TINGGI"
        elif vol_ratio < 0.8:
            volume_status = "RENDAH"
        else:
            volume_status = "NORMAL"
        
        # Bollinger Band position
        if current_price < bb_lower:
            bb_position = "DI BAWAH LOWER BAND"
        elif current_price > bb_upper:
            bb_position = "DI ATAS UPPER BAND"
        elif current_price > bb_middle:
            bb_position = "UPPER HALF"
        else:
            bb_position = "LOWER HALF"
        
        # Stochastic
        if stoch_k < 20:
            stoch_status = "OVERSOLD"
        elif stoch_k > 80:
            stoch_status = "OVERBOUGHT"
        else:
            stoch_status = "NORMAL"
        
        # ADX - Trend strength
        if adx > 25:
            adx_status = "TREND KUAT"
        elif adx > 15:
            adx_status = "TREND MODERATE"
        else:
            adx_status = "TREND LEMAH"
        
        # Market condition
        market_condition = detect_market_condition(df, ma50, ma200)
        
        # Data quality score
        data_quality = min(100, (len(df) / 252) * 100) if len(df) < 252 else 100
        
        return {
            "rsi": rsi,
            "rsi_status": rsi_status,
            "macd": macd_val,
            "macd_signal": macd_sig,
            "macd_histogram": macd_hist,
            "macd_status": macd_status,
            "macd_hist_status": macd_hist_status,
            "ma50": ma50,
            "ma200": ma200,
            "ma20": ma20,
            "trend_status": trend_status,
            "volume_current": round(current_vol, 2),
            "volume_avg": round(avg_vol, 2),
            "volume_ratio": vol_ratio,
            "volume_status": volume_status,
            "support": support,
            "resistance": resistance,
            "current_price": current_price,
            "market_condition": market_condition,
            # Advanced indicators
            "bb_upper": bb_upper,
            "bb_middle": bb_middle,
            "bb_lower": bb_lower,
            "bb_position": bb_position,
            "atr": atr,
            "stoch_k": stoch_k,
            "stoch_d": stoch_d,
            "stoch_status": stoch_status,
            "adx": adx,
            "adx_status": adx_status,
            "vwap": vwap,
            # Fibonacci & Pivot
            "fib_382": fib_382,
            "fib_500": fib_500,
            "fib_618": fib_618,
            "pivot": pivot,
            "r1": r1,
            "s1": s1,
            "r2": r2,
            "s2": s2,
            # Volume Flow - OBV
            "obv": round(obv, 2) if obv is not None else 0,
            "obv_prev": round(obv_prev, 2) if obv_prev is not None else 0,
            "obv_divergence": obv_divergence,
            # Data quality
            "data_quality": data_quality
        }
    except Exception as e:
        print(f"⚠️ Error stock indicators: {e}")
        return None
        return None

# ==============================
# GOOGLE SHEETS CLIENT
# ==============================
def get_sheets_client():
    try:
        creds_json = os.getenv("GOOGLE_CREDENTIALS")
        if creds_json:
            creds_dict = json.loads(creds_json)
        else:
            with open("credentials.json", "r") as f:
                creds_dict = json.load(f)
                
        scopes = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        gc = gspread.authorize(creds)
        return gc
    except Exception as e:
        print(f"❌ Google Sheets error: {e}")
        return None

# ==============================
# 1. ADAPTIVE LEARNING
# ==============================
def get_trade_history():
    """Ambil histori trade dari sheet Performance untuk adaptive learning"""
    try:
        gc = get_sheets_client()
        if not gc:
            return []
        
        sheet_id = os.getenv("GOOGLE_SHEETS_ID")
        spreadsheet = gc.open_by_key(sheet_id)
        ws = spreadsheet.worksheet("Performance")
        
        all_values = ws.get_all_values()
        if len(all_values) < 2:
            return []
        
        # Cari header row
        header_idx = 0
        for i, row in enumerate(all_values):
            if "Tanggal" in row or "Status" in row:
                header_idx = i
                break
        
        headers = all_values[header_idx]
        data_rows = all_values[header_idx + 1:]
        
        trades = []
        for row in data_rows:
            if len(row) >= len(headers):
                trade = dict(zip(headers, row))
            elif len(row) > 0 and row[0]:
                padded = row + [""] * (len(headers) - len(row))
                trade = dict(zip(headers, padded))
            else:
                continue
            
            # Parse skor_detail jika ada
            skor_json = trade.get("Skor Detail", "")
            if skor_json:
                try:
                    trade["skor_detail"] = json.loads(skor_json)
                except:
                    trade["skor_detail"] = None
            else:
                trade["skor_detail"] = None
            
            # Normalize
            trade["hasil"] = trade.get("Hasil", "")
            trade["status"] = trade.get("Status", "")
            
            if trade.get("status") == "CLOSED":
                trades.append(trade)
        
        return trades
    except Exception as e:
        print(f"⚠️ Error ambil trade history: {e}")
        return []

def calculate_adaptive_weights(trade_history):
    """
    Jika histori >= 30 trade CLOSED dengan skor_detail:
    - Evaluasi win rate tiap indikator
    - Naikkan bobot indikator dengan win rate tinggi
    - Turunkan bobot indikator dengan win rate rendah
    """
    weights = DEFAULT_WEIGHTS.copy()
    
    # Filter trades yang punya skor_detail
    trades_with_scores = [t for t in trade_history if t.get("skor_detail")]
    
    if len(trades_with_scores) < 30:
        return weights
    
    for indicator in ["rsi", "macd", "trend", "volume", "sentimen"]:
        wins = 0
        total = 0
        
        for trade in trades_with_scores:
            scores = trade["skor_detail"]
            if indicator in scores and scores[indicator] != 0:
                total += 1
                if trade.get("hasil") == "WIN":
                    wins += 1
        
        if total >= 10:
            win_rate = wins / total
            if win_rate > 0.65:
                weights[indicator] = 1.5  # Naikkan bobot (+50%)
            elif win_rate > 0.55:
                weights[indicator] = 1.2  # Sedikit naik
            elif win_rate < 0.35:
                weights[indicator] = 0.5  # Turunkan bobot (-50%)
            elif win_rate < 0.45:
                weights[indicator] = 0.8  # Sedikit turun
    
    return weights

# ==============================
# 4. ANALISIS BERITA (IMPROVED)
# ==============================
def analyze_news(berita, symbol):
    """
    Analisis berita dengan label: LANGSUNG / SEKTOR / MAKRO
    Ambil maksimal 3 berita, tentukan sentimen, skor, dan dampak
    """
    if not berita:
        return {
            "berita_label": [],
            "status": "NETRAL",
            "skor": 0.0,
            "dampak": "Tidak ada berita signifikan"
        }
    
    # Ambil maksimal 3 berita
    top_berita = berita[:3]
    headlines = "\n".join([f"- {b['judul']} ({b.get('sumber', 'N/A')})" for b in top_berita])
    
    # Deteksi label berita (LANGSUNG/SEKTOR/MAKRO)
    symbol_upper = symbol.upper()
    is_crypto = not symbol.endswith(".JK")
    
    labeled_berita = []
    has_makro_only = True
    for b in top_berita:
        judul = b.get('judul', '').lower()
        label = "LANGSUNG"
        if is_crypto:
            if any(x in judul for x in ['sector', 'industry', 'mining', 'energy', 'bank', 'tech']):
                label = "SEKTOR"
            if any(x in judul for x in ['fed', 'interest rate', 'inflation', 'gdp', 'economy', 'recession', 'employment']):
                label = "MAKRO"
        else:
            if symbol_upper in judul:
                label = "LANGSUNG"
            elif any(x in judul for x in ['sector', 'industry', 'banking', 'tech']):
                label = "SEKTOR"
            else:
                label = "MAKRO"
        
        if label != "MAKRO":
            has_makro_only = False
        
        labeled_berita.append({
            "judul": b.get('judul', ''),
            "sumber": b.get('sumber', 'N/A'),
            "label": label
        })
    
    # Deteksi berita BESAR yang mempengaruhi sentimen
    big_news_keywords = [' wafat', 'meninggal', 'died', 'passed away', 'ceo resign', 'resign', 
                       'peraturan', 'regulation', 'ban', 'prohibited', 'geopolitik', 'war',
                       'scandal', 'fraud', 'investigation', 'akuisisi', 'merger']
    
    has_big_news = any(any(kw in b.get('judul', '').lower() for kw in big_news_keywords) for b in top_berita)
    
    # Override dampak jika hanya berita makro
    makro_override = "Dampak tidak langsung ke saham" if has_makro_only else None
    
    prompt = f"""Analisis sentimen berita untuk aset {symbol}.

Berita:
{headlines}

JAWAB HANYA dalam format JSON (tanpa markdown, tanpa penjelasan tambahan):
{{"status": "POSITIF/NETRAL/NEGATIF", "skor": <float -1.0 sampai 1.0>, "dampak": "<dampak ke harga dalam 1 kalimat>"}}"""
    
    try:
        response = client.chat.completions.create(
            model="meta-llama/llama-3.3-70b-instruct",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}]
        )
        result = response.choices[0].message.content.strip()
        
        if "```" in result:
            parts = result.split("```")
            for part in parts:
                part = part.strip()
                if part.startswith("json"):
                    part = part[4:].strip()
                if part.startswith("{"):
                    result = part
                    break
        
        json_match = re.search(r'\{[^}]+\}', result)
        if json_match:
            parsed = json.loads(json_match.group())
            status = parsed.get("status", "NETRAL").upper()
            if status not in ["POSITIF", "NETRAL", "NEGATIF"]:
                status = "NETRAL"
            skor = max(-1.0, min(1.0, float(parsed.get("skor", 0))))
            dampak = str(parsed.get("dampak", "Tidak ada dampak signifikan"))
            
            # Override jika hanya berita makro
            if makro_override:
                dampak = makro_override
            
            return {
                "berita_label": labeled_berita,
                "status": status,
                "skor": skor,
                "dampak": dampak
            }
        
        return {
            "berita_label": labeled_berita,
            "status": "NETRAL",
            "skor": 0.0,
            "dampak": makro_override if makro_override else "Tidak ada sentimen signifikan"
        }
    except Exception as e:
        print(f"⚠️ Error sentiment analysis: {e}")
        return {
            "berita_label": labeled_berita,
            "status": "NETRAL",
            "skor": 0.0,
            "dampak": makro_override if makro_override else "Tidak ada sentimen signifikan"
        }

# ==============================
# 5. SCORING ENGINE
# ==============================
def calculate_score(indikator, sentimen, weights, market_condition, trend_direction):
    """
    Hitung skor deterministik dari semua indikator dengan logika contextual.

    RSI: Contextual berdasarkan trend_direction
    Stochastic: Contextual berdasarkan trend_direction
    MACD: +2 / -2
    Trend: +3 / 0 / -3
    Volume: +1 / -1
    Sentimen: +2 / 0 / -2

    Total skor = jumlah semua × bobot adaptif
    """
    scores = {}

    # === CONTEXTUAL RSI & STOCHASTIC SCORING ===
    rsi = indikator.get("rsi", 50)
    rsi_status = indikator.get("rsi_status", "NORMAL")
    stoch_k = indikator.get("stoch_k", 50)
    stoch_status = indikator.get("stoch_status", "NORMAL")

    # Determine if strong trend
    is_strong_bullish = trend_direction == "BULLISH (dominan)"
    is_strong_bearish = trend_direction == "BEARISH (dominan)"
    is_sideways = market_condition == "SIDEWAYS"

    # RSI Contextual Scoring (Overhaul)
    if trend_direction == "BULLISH" and rsi > 70:
        scores["rsi"] = 1  # Overbought = Strength in uptrend
    elif market_condition == "SIDEWAYS" and rsi > 70:
        scores["rsi"] = -2  # Overbought = Reversal Risk in range
    elif rsi_status == "OVERSOLD":
        scores["rsi"] = 2
    elif rsi_status == "OVERBOUGHT":
        scores["rsi"] = -2
    else:
        scores["rsi"] = 0

    # Stochastic Contextual Scoring
    if is_strong_bullish:
        # Dalam trend bullish kuat: Overbought adalah STRENGTH
        if stoch_k > 80 or stoch_status == "OVERBOUGHT":
            scores["stoch"] = 1  # Strength
        elif stoch_k < 30 or stoch_status == "OVERSOLD":
            scores["stoch"] = -1  # Weakness
        else:
            scores["stoch"] = 0
    elif is_strong_bearish:
        # Dalam trend bearish kuat: Stoch 40-50 adalah RESISTANCE
        if stoch_k >= 40 and stoch_k <= 50:
            scores["stoch"] = -1  # Resistance
        elif stoch_k < 25 or stoch_status == "OVERSOLD":
            scores["stoch"] = 0  # Normal weakness
        elif stoch_k > 75 or stoch_status == "OVERBOUGHT":
            scores["stoch"] = 1  # Unexpected strength
        else:
            scores["stoch"] = 0
    elif is_sideways:
        # Sideways: Standard oscillator rules
        if stoch_status == "OVERSOLD":
            scores["stoch"] = 2
        elif stoch_status == "OVERBOUGHT":
            scores["stoch"] = -2
        else:
            scores["stoch"] = 0
    else:
        # Default: Standard rules
        if stoch_status == "OVERSOLD":
            scores["stoch"] = 2
        elif stoch_status == "OVERBOUGHT":
            scores["stoch"] = -2
        else:
            scores["stoch"] = 0
    
    # MACD Score
    macd_status = indikator.get("macd_status", "BEARISH")
    scores["macd"] = 2 if macd_status == "BULLISH" else -2
    
    # Trend Score (PRICE vs MA - STRICT LOGIC)
    # Prioritas: Price Position (di bawah/diatas semua MA) > ADX > trend_status
    current_price = indikator.get("current_price", 0)
    ma20 = indikator.get("ma20", 0)
    ma50 = indikator.get("ma50", 0)
    ma200 = indikator.get("ma200", 0)
    adx = indikator.get("adx", 0)
    
    if current_price > 0 and ma20 > 0 and ma50 > 0 and ma200 > 0:
        # PRIORITAS TERTINGGI: Harga di bawah semua MA = mutlak bearish
        if current_price < ma20 and current_price < ma50 and current_price < ma200:
            scores["trend"] = -3  # Mutlak bearish (Prioritas Tertinggi)
        # PRIORITAS TERTINGGI: Harga di atas semua MA = mutlak bullish
        elif current_price > ma20 and current_price > ma50 and current_price > ma200:
            scores["trend"] = 3   # Mutlak bullish (Prioritas Tertinggi)
        # ADX < 20 HANYA jika harga tersangkut di antara garis MA
        elif adx > 0 and adx < 20:
            scores["trend"] = 0   # Sideways HANYA jika harga tersangkut di antara garis MA
        else:
            trend_status = indikator.get("trend_status", "NEUTRAL")
            scores["trend"] = 3 if trend_status == "BULLISH" else -3 if trend_status == "BEARISH" else 0
    else:
        trend_status = indikator.get("trend_status", "NEUTRAL")
        scores["trend"] = 3 if trend_status == "BULLISH" else -3 if trend_status == "BEARISH" else 0
    
    # Volume Score — SMART LOGIC
    volume_status = indikator.get("volume_status", "NORMAL")
    direction = scores["macd"] + scores["trend"]
    
    if volume_status == "TINGGI":
        if direction > 0:
            scores["volume"] = 1   # Konfirmasi searah trend
        elif direction < 0:
            scores["volume"] = 0   # Volume tinggi tapi lawan trend
        else:
            scores["volume"] = 0
    elif volume_status == "RENDAH":
        scores["volume"] = -1  # Volume lemah = sinyal lemah
    else:
        scores["volume"] = 0
    
    # Sentimen Score (Strict Variable Binding)
    scores["sentimen"] = float(sentimen.get("skor", 0))
    
    # === TREND DOMINANCE RULE ===
    # Jika indikator melawan trend dengan volume tinggi → kurangi skor -1
    is_trending_down = market_condition.startswith("TRENDING DOWN")
    is_trending_up = market_condition.startswith("TRENDING UP")
    
    if is_trending_down and volume_status == "TINGGI":
        if scores["macd"] == 2:  # Bullish tapi trend turun
            scores["macd"] = max(0, scores["macd"] - 1)
        if scores["rsi"] == 2:   # Oversold tapi trend turun
            scores["rsi"] = max(0, scores["rsi"] - 1)
    
    if is_trending_up and volume_status == "TINGGI":
        if scores["macd"] == -2:  # Bearish tapi trend naik
            scores["macd"] = min(0, scores["macd"] + 1)
        if scores["rsi"] == -2:  # Overbought tapi trend naik
            scores["rsi"] = min(0, scores["rsi"] + 1)
    
    # === APPLY ADAPTIVE WEIGHTS ===
    total = 0.0
    for key in scores:
        w = weights.get(key, 1.0)
        total += scores[key] * w
    
    return scores, round(total, 2)

# ==============================
# 8. NO TRADE ZONE & BREAKOUT DETECTION
# ==============================
def detect_no_trade_zone(indikator, skor_detail):
    """
    Jika sideways + volume lemah + konflik indikator → Hindari entry
    """
    market = indikator.get("market_condition", "")
    vol_status = indikator.get("volume_status", "")
    
    # Cek konflik indikator (ada yang positif dan negatif)
    positive_count = sum(1 for v in skor_detail.values() if v > 0)
    negative_count = sum(1 for v in skor_detail.values() if v < 0)
    has_conflict = positive_count > 0 and negative_count > 0
    
    if market == "SIDEWAYS" and vol_status == "RENDAH" and has_conflict:
        return True
    
    return False


def detect_breakout_breakdown(indikator, market_condition):
    """
    Deteksi breakout (buy confirmation) atau breakdown (sell confirmation)
    
    BREAKOUT (BUY confirmation):
    - Harga menembus resistance + volume tinggi
    
    BREAKDOWN (SELL confirmation):
    - Harga menembus support + volume tinggi
    
    Returns: "breakout", "breakdown", atau "none"
    """
    current_price = indikator.get("current_price", 0)
    support = indikator.get("support", 0)
    resistance = indikator.get("resistance", 0)
    volume_status = indikator.get("volume_status", "NORMAL")
    vol_ratio = indikator.get("volume_ratio", 1.0)
    
    if current_price <= 0 or support <= 0 or resistance <= 0:
        return "none"
    
    # Buffer untuk breakout/breakdown (2% dari S/R)
    breakout_threshold = resistance * 1.02
    breakdown_threshold = support * 0.98
    
    # Volume tinggi atau meningkat (>1.2x rata-rata)
    is_volume_confirmed = volume_status == "TINGGI" or vol_ratio > 1.2
    
    # Deteksi breakout (harga di atas resistance dengan volume)
    if current_price >= breakout_threshold and is_volume_confirmed:
        return "breakout"
    
    # Deteksi breakdown (harga di bawah support dengan volume)
    if current_price <= breakdown_threshold and is_volume_confirmed:
        return "breakdown"
    
    return "none"

# ==============================
# 9. ENTRY / SL / TP CALCULATION (DYNAMIC RISK MANAGEMENT)
# ==============================
def calculate_entry_sl_tp(harga, sinyal, indikator):
    support = indikator.get("support", harga * 0.97)
    resistance = indikator.get("resistance", harga * 1.03)
    r1 = indikator.get("r1", resistance)
    s1 = indikator.get("s1", support)
    atr = indikator.get("atr", 0)
    
    is_buy = "BUY" in sinyal.upper()
    is_sell = "SELL" in sinyal.upper()
    
    if not is_buy and not is_sell:
        return 0, 0, 0
    
    entry = harga
    sl = 0
    tp = 0
    
    if is_buy:
        atr_sl = harga - (1.5 * atr) if atr > 0 else 0
        support_sl = support * 0.98
        sl = max(atr_sl, support_sl) if atr_sl > 0 else support_sl
        
        tp = r1 if r1 > entry else resistance
        if tp <= entry: 
            tp = entry * 1.05
            
    elif is_sell:
        atr_sl = harga + (1.5 * atr) if atr > 0 else float('inf')
        resistance_sl = resistance * 1.02
        sl = min(atr_sl, resistance_sl) if atr_sl != float('inf') else resistance_sl
        
        tp = s1 if s1 < entry else support
        if tp >= entry: 
            tp = entry * 0.95
            
    return round(entry, 4), round(sl, 4), round(tp, 4)

# ==============================
# 10. COPY TRADING — POSITION SIZING
# ==============================
def calculate_position_size(modal, entry, sl, sinyal, market_condition="TRENDING", currency="USD"):
    """
    Position Size = (Risk% × Modal) / |Entry - Stop Loss|
    Risk: 1% normal, 0.5% volatile
    Konversi modal ke mata uang yang sama dengan entry price
    """
    if sinyal == "HOLD" or entry == 0:
        return 0
    
    # Risk per trade: 1% normal, 0.5% kalau volatile
    is_volatile = "VOLATILE" in market_condition
    risk_pct = 0.005 if is_volatile else 0.01
    
    # Konversi modal ke mata uang yang sama dengan entry
    if currency == "USD":
        # Modal default IDR, convert ke USD (approx rate 16000 IDR/USD)
        modal_usd = modal / 16000
        risk_per_trade = modal_usd * risk_pct
    elif currency == "IDR":
        risk_per_trade = modal * risk_pct
    else:
        risk_per_trade = modal * risk_pct
    
    risk_per_unit = abs(entry - sl)
    
    if risk_per_unit == 0:
        return 0
    
    size = risk_per_trade / risk_per_unit
    return round(size, 6)


# ==============================
# 11. PROFESSIONAL RISK METRICS
# ==============================
def calculate_risk_metrics(harga, entry, sl, tp, indikator, sinyal, modal, currency="USD"):
    """
    Calculate professional risk metrics:
    - Risk/Reward Ratio (RR)
    - Kelly Criterion (win rate estimation)
    - Risk per trade (Rp/USD)
    - Position value
    - Max risk exposure
    """
    metrics = {}
    
    is_buy = "BUY" in sinyal
    is_sell = "SELL" in sinyal
    
    if not is_buy and not is_sell:
        return {
            "rr_ratio": "-",
            "risk_reward": 0,
            "risk_amount": 0,
            "reward_amount": 0,
            "position_value": 0,
            "kelly_pct": 0,
            "max_risk_pct": 1.0,
            "is_valid_rr": False
        }
    
    # Risk and Reward calculation
    risk_pct = abs(entry - sl) / entry * 100
    
    if is_buy:
        reward_pct = abs(tp - entry) / entry * 100
    else:  # SELL
        reward_pct = abs(entry - tp) / entry * 100
    
    # RR Ratio
    if risk_pct > 0:
        rr = reward_pct / risk_pct
        rr_ratio = f"1:{rr:.1f}"
        is_valid_rr = rr >= 1.5
    else:
        rr = 0
        rr_ratio = "N/A"
        is_valid_rr = False
    
    # Risk and Reward in currency
    risk_amount = abs(entry - sl)
    reward_amount = abs(tp - entry)
    
    # Position value
    if currency == "USD":
        risk_amt_idr = risk_amount * 16000
        position_value = modal
    else:
        risk_amt_idr = risk_amount
        position_value = modal
    
    # Kelly Criterion (simplified - assumes 50% base win rate, adjusted by RR)
    # Kelly % = WinRate - (LossRate / RR)
    # For conservative estimate, assume 50% win rate
    win_rate = 0.5
    kelly_pct = win_rate - ((1 - win_rate) / rr) if rr > 0 else 0
    kelly_pct = max(0, min(100, kelly_pct * 100))
    
    # Max risk exposure (1% of capital)
    max_risk_pct = 1.0
    
    # Calculate position size for Kelly
    risk_per_trade_pct = risk_pct
    
    return {
        "rr_ratio": rr_ratio,
        "risk_reward": round(rr, 2),
        "risk_pct": round(risk_pct, 2),
        "reward_pct": round(reward_pct, 2),
        "risk_amount": round(risk_amount, 4),
        "reward_amount": round(reward_amount, 4),
        "position_value": position_value,
        "kelly_pct": round(kelly_pct, 1),
        "max_risk_pct": max_risk_pct,
        "is_valid_rr": is_valid_rr,
        "risk_per_trade_pct": round(risk_per_trade_pct, 2)
    }


# ==============================
# 12. DATA QUALITY SCORE
# ==============================
def calculate_data_quality(indikator, berita):
    """
    Calculate data quality score based on:
    - Indicator data completeness
    - News availability
    - Data freshness
    - Data anomaly detection (MA flat, etc.)
    """
    score = 0
    details = []
    warnings = []
    
    # 1. Indicator quality (50%)
    if indikator:
        indicator_score = 50
        
        # Check data quality from indicator
        dq = indikator.get("data_quality", 0)
        if dq > 0:
            indicator_score = dq / 2  # Max 50%
        
        # Additional checks
        if indikator.get("ma50") and indikator.get("ma200"):
            indicator_score += 10
        if indikator.get("atr", 0) > 0:
            indicator_score += 10
        if indikator.get("adx", 0) > 0:
            indicator_score += 10
        
        indicator_score = round(min(80, indicator_score), 1)
        score += indicator_score
        details.append(f"Indikator: {indicator_score}%")
        
        # === DATA ANOMALY CHECK ===
        
        # Check MA flat (MA20 ≈ MA50)
        ma20 = indikator.get("ma20", 0)
        ma50 = indikator.get("ma50", 0)
        if ma20 and ma50 and ma20 != 0 and ma50 != 0:
            ma_diff_pct = abs(ma20 - ma50) / ma50 * 100
            if ma_diff_pct < 0.5:  # Less than 0.5% difference
                warnings.append("MA20 ≈ MA50 (flat trend)")
                score = max(0, score - 15)
        
        # Check price position vs MA
        current_price = indikator.get("current_price", 0)
        ma200 = indikator.get("ma200", 0)
        if current_price and ma50 and ma200 and ma200 != 0:
            if current_price < ma50 and current_price < ma200:
                # Already bearish, verify trend consistency
                pass
            elif current_price > ma50 and current_price > ma200:
                # Already bullish, verify trend consistency
                pass
        
        # Check if ADX exists but trend is unclear
        adx = indikator.get("adx", 0)
        trend_status = indikator.get("trend_status", "NEUTRAL")
        if adx > 25 and trend_status == "NEUTRAL":
            warnings.append("ADX tinggi tapi trend NETRAL (inkonsistensi)")
            score = max(0, score - 10)
    
    # 2. News quality (30%)
    if berita:
        news_score = min(30, len(berita) * 10)
        score += news_score
        details.append(f"Berita: {round(news_score, 1)}%")
    else:
        details.append("Berita: 0%")
    
    # 3. Additional factors (20%)
    if indikator:
        if indikator.get("volume_ratio", 1) > 0:
            score += 10
        if indikator.get("current_price", 0) > 0:
            score += 10
    
    score = round(min(100, score), 1)
    
    # Determine grade with warning
    grade = "A" if score >= 80 else "B" if score >= 60 else "C" if score >= 40 else "D"
    if warnings:
        grade = f"WARNING ({grade})"
    
    return {
        "quality_score": score,
        "quality_grade": grade,
        "details": ", ".join(details),
        "warnings": warnings
    }

# ==============================
# AI REASONING (LLM untuk narasi final)
# ==============================
def get_ai_reasoning(symbol, indikator, sentimen, skor_detail, total_skor, sinyal, market_condition, confidence, rr_ratio, risk_level, execution_status, edge_clarity, next_trade_plan):
    try:
        rsi = float(indikator.get("rsi", 50))
        adx = float(indikator.get("adx", 0))
        stoch_str = str(indikator.get("stochastic", "")).upper()
        vol_str = str(indikator.get("volume", "")).upper()
        obv_str = str(indikator.get("obv_flow", "")).upper()
        current_price = float(indikator.get("current_price", 0))
        resistance = float(indikator.get("resistance", 0))
    except Exception:
        rsi, adx, stoch_str, vol_str, obv_str, current_price, resistance = 50, 0, "", "", "", 0, 0

    # Exact Terminology Validation
    if confidence >= 60:
        nama_sinyal_tabel = "READY (Waiting Trigger)"
    elif confidence >= 40:
        nama_sinyal_tabel = "WATCHLIST (Active Monitoring)"
    elif confidence >= 30:
        nama_sinyal_tabel = "LOW PRIO WATCHLIST (No Man's Land)"
    else:
        nama_sinyal_tabel = "IGNORE (No Valid Setup)"

    # S/R Contextual Awareness
    is_testing_res = False
    if resistance > 0 and current_price > 0:
        if abs(resistance - current_price) / resistance * 100 <= 2.0:
            is_testing_res = True
            nama_sinyal_tabel = nama_sinyal_tabel.replace("No Man's Land", "Menguji Resistance")
            if "No Man's Land" in edge_clarity:
                edge_clarity = edge_clarity.replace("No Man's Land", "Menguji Resistance")

    adx_desc = "Sangat Kuat" if adx >= 40 else ("Kuat" if adx >= 25 else ("Moderate/Mulai Terbentuk" if adx >= 20 else "Lemah/Choppy"))

    kondisi_harga = ""
    if adx < 20: 
        if total_skor >= 5 and "SELL" in next_trade_plan:
            kondisi_harga = f"PARADOKS: Skor Makro Bullish (+{total_skor}), namun market Sideways dan harga di Resistance. Strategi dikalibrasi ke Counter-Trend Short (Range Sell)."
        elif total_skor <= -5 and "BUY" in next_trade_plan:
            kondisi_harga = f"PARADOKS: Skor Makro Bearish ({total_skor}), namun market Sideways dan harga di Support. Strategi dikalibrasi ke Counter-Trend Long (Range Buy)."
        elif "DOWN" in market_condition and "DISTRIBUTION" in obv_str:
            kondisi_harga = "Market Sideways dengan tekanan Bearish dominan (OBV Distribusi). Peluang terbaik adalah SELL di resistance."
        elif "UP" in market_condition and "ACCUMULATION" in obv_str:
            kondisi_harga = "Market Sideways dengan sokongan Bullish (OBV Akumulasi). Peluang terbaik adalah BUY di support."
    else: 
        if "DOWN" in market_condition:
            if rsi >= 70 or "OVERBOUGHT" in stoch_str:
                kondisi_harga = "Harga mengalami Counter-Trend Rally (Overbought). Ini memperkuat probabilitas setup SELL ON RALLY di resistance."
            elif total_skor > -3:
                kondisi_harga = "Tren utama Bearish, namun momentum jangka pendek sedang naik. Ini adalah fase PULLBACK (Counter-Trend Rally), momentum ideal untuk mencari pijakan SELL ON RALLY di area resistance."
        elif "UP" in market_condition:
            if rsi <= 30 or "OVERSOLD" in stoch_str:
                kondisi_harga = "Harga terkoreksi (Oversold). Ini memperkuat probabilitas setup BUY ON DIP di support."
            elif total_skor < 3:
                kondisi_harga = "Tren utama Bullish, namun momentum jangka pendek melemah. Ini adalah fase KOREKSI, momentum ideal untuk mencari pijakan BUY ON DIP di area support."
            
    if is_testing_res:
        kondisi_harga += " PERHATIAN: Harga saat ini sedang menguji Resistance (Rentan Rejection)."
        
    vol_context = "Volume transaksi yang sepi menandakan market belum siap breakout." if "RENDAH" in vol_str or ("NORMAL" in vol_str and adx < 20) else ""
        
    if confidence == 0 or "IGNORE" in sinyal.upper() or "HOLD" in sinyal.upper():
        branching_rules = "2. ANTI-HALLUCINATION WAJIB: Status adalah NO TRADE atau IGNORE. Narasi HANYA boleh menjelaskan mengapa kriteria tidak terpenuhi. DILARANG KERAS menyebutkan angka Risk/Reward (R:R), persentase keuntungan, kalimat penjelasan R:R, Take Profit, atau Stop Loss. Dilarang memunculkan frasa kontradiktif seperti 'System Confidence tinggi'."
    elif confidence < 50:
        branching_rules = "2. JIKA Confidence < 50%: WAJIB cantumkan kalimat: 'Sistem belum memiliki keyakinan yang cukup untuk eksekusi, sehingga aksi hanya sebatas pemantauan (Watchlist)'. DILARANG KERAS menggunakan frasa 'keyakinan yang cukup untuk melakukan aksi'."
    elif confidence < 60:
        branching_rules = "2. JIKA Confidence < 60%: DILARANG KERAS menggunakan frasa 'keyakinan yang cukup untuk melakukan aksi' karena status belum READY/EXECUTE."
    else:
        branching_rules = "2. JIKA Confidence > 60%: WAJIB berikan kalimat penjelas eksplisit di akhir paragraf mengapa angka tersebut tinggi (berdasarkan konfluensi indikator, dll)."

    prompt = f"""Sebagai AI Quant Trader tingkat lanjut, buat laporan eksekusi untuk {symbol}.

[DATA KUANTITATIF]
Tren Makro: {market_condition} | Kekuatan Tren: {adx_desc} (ADX: {adx:.1f})
Analisis Mendalam: {kondisi_harga} {vol_context}
Status Sistem (Edge Clarity): {edge_clarity} | System Confidence: {confidence}%
Sinyal Internal: {sinyal}
Nama Sinyal Tampilan: {nama_sinyal_tabel}

[ACTION PLAN LENGKAP]
{next_trade_plan}

ATURAN MUTLAK PENULISAN (ZERO TOLERANCE):
1. TATA BAHASA: Gunakan Bahasa Indonesia baku. HANYA gunakan alfabet Latin (A-Z). DILARANG KERAS memakai aksara Sirilik (Cyrillic), Arab, atau encoding bahasa asing apapun! Nama sinyal dalam teks WAJIB 100% sama dengan "{nama_sinyal_tabel}". JANGAN diubah menjadi frasa lain (seperti "Neutral Watch").
{branching_rules}
3. JIKA Status bukan NO TRADE: Anda WAJIB mengutip format Action Plan secara LENGKAP (Strategi, TP, SL, R:R). 
4. KALIMAT PERTAMA WAJIB langsung mengutip isi dari 'Analisis Mendalam' (jika ada) atau status Tren makro. JANGAN PERNAH MEMOTONG KALIMAT DI TENGAH JALAN.
"""

    try:
        response = client.chat.completions.create(
            model="meta-llama/llama-3.3-70b-instruct",
            max_tokens=600, # Diperbesar agar LLM tidak kehabisan napas
            messages=[{"role": "user", "content": prompt}]
        )
        result = response.choices[0].message.content
        # Proteksi keamanan ganda jika result terpotong
        if not result.endswith((".", "!", "?")):
            result += f"... [Teks terpotong]. Detail Plan: {next_trade_plan}"
            
        return result.strip() if result else f"[Sistem Otomatis]\n{next_trade_plan}"
    except Exception as e:
        print(f"⚠️ LLM Error: {e}")
        return f"⚠️ LLM Sedang Sibuk. Raw Plan: {next_trade_plan}"

# ==============================
# FORMAT OUTPUT
# ==============================
def fmt_price(price, currency="USD"):
    """Format harga sesuai currency"""
    if price is None:
        return "N/A"
    if currency == "IDR":
        return f"Rp {price:,.0f}"
    elif price >= 1:
        return f"${price:,.2f}"
    else:
        return f"${price:,.6f}"

def format_analysis_output(symbol, harga, harga_idr, indikator, sentimen,
                            sinyal, total_skor, confidence, entry, sl, tp,
                            risk_level, rr_ratio, modal, position_size, alasan,
                            market_condition, skor_detail, weights, jenis, no_trade,
                            is_early_entry=False, risk_metrics=None, data_quality=None):
    """Format output sesuai template 10-point system"""

    # --- EARLY GLOBAL PATCH (UI SYNC) ---
    # 1. OBV Patch (Mencegah teks NONE muncul di laporan)
    try:
        obv_val = float(indikator.get("obv", 0))
        if indikator.get("obv_flow", "NONE") in ["NONE", ""]:
            indikator["obv_flow"] = "BULLISH (Accumulation)" if obv_val > 0 else "BEARISH (Distribution)"
    except Exception:
        pass

    # Catatan untuk AI IDE: Pastikan nilai indikator["obv_flow"] yang baru ini
    # yang dimasukkan ke dalam f-string laporan teknikal!

    # 2. Sync Sinyal & Risk Level (Dijalankan SETELAH kalkulasi Confidence dari Engine AI selesai)
    if confidence >= 60:
        sinyal = "READY (Waiting Trigger)"
        risk_level = "MEDIUM (Actionable Setup)"
        copy_trade_status = "PENDING EXECUTION (Waiting Trigger)"
    elif confidence >= 40:
        sinyal = "WATCHLIST (Active Monitoring)"
        risk_level = "LOW-MEDIUM (Setup Developing)"
        copy_trade_status = "CONDITIONAL WATCHLIST"
    elif confidence >= 30:
        sinyal = "LOW PRIO WATCHLIST (No Man's Land)"
        risk_level = "LOW (Passive Monitoring)"
        copy_trade_status = "NO TRADE"
    else:
        sinyal = "IGNORE (No Valid Setup)"
        risk_level = "LOW (Avoid)"
        copy_trade_status = "NO TRADE"
    # ------------------------------------

    # Default values
    if risk_metrics is None:
        risk_metrics = {
            "rr_ratio": "-", "risk_reward": 0, "risk_pct": 0, "reward_pct": 0,
            "kelly_pct": 0, "is_valid_rr": False
        }
    if data_quality is None:
        data_quality = {"quality_score": 50, "quality_grade": "C", "details": "Data tidak tersedia"}
    
    # Currency tergantung jenis aset
    if jenis == "Crypto":
        curr = "USD"
        harga_display = fmt_price(harga, "USD")
        if harga_idr:
            harga_display += f" (Rp {harga_idr:,.0f})"
        entry_display = fmt_price(entry, "USD")
        sl_display = fmt_price(sl, "USD")
        tp_display = fmt_price(tp, "USD")
    else:
        # Saham Indonesia pakai IDR, US pakai USD
        if symbol.endswith(".JK"):
            curr = "IDR"
        else:
            curr = "USD"
        harga_display = fmt_price(harga, curr)
        entry_display = fmt_price(entry, curr)
        sl_display = fmt_price(sl, curr)
        tp_display = fmt_price(tp, curr)
    
# Support/Resistance display
    s_display = fmt_price(indikator.get("support"), curr)
    r_display = fmt_price(indikator.get("resistance"), curr)
    
    # MA display - MA20, MA50, MA200
    ma20_str = fmt_price(indikator.get("ma20"), curr) if indikator.get("ma20") else "N/A"
    ma50_str = fmt_price(indikator.get("ma50"), curr) if indikator.get("ma50") else "N/A"
    ma200_str = fmt_price(indikator.get("ma200"), curr) if indikator.get("ma200") else "N/A"
    
    # Adaptive weights info
    weights_modified = any(w != 1.0 for w in weights.values())
    weights_note = ""
    if weights_modified:
        mods = []
        for k, v in weights.items():
            if v != 1.0:
                mods.append(f"{k.upper()}: {v}x")
        weights_note = f"\n📐 Bobot Adaptif: {', '.join(mods)}"
    
    # No-trade warning
    no_trade_warning = ""
    if no_trade:
        no_trade_warning = "\n⛔ NO TRADE ZONE: Sideways + Volume lemah + Konflik indikator → Hindari entry!"

    # HOLD dengan bias
    if "HOLD" in sinyal.upper():
        if total_skor > 0:
            sinyal_display = "HOLD (Bullish Bias)"
        elif total_skor < 0:
            sinyal_display = "HOLD (Bearish Bias)"
        else:
            # Skor = 0 → tentukan dari trend dominan
            if "TRENDING UP" in market_condition:
                sinyal_display = "HOLD (Bullish Bias)"
            elif "TRENDING DOWN" in market_condition:
                sinyal_display = "HOLD (Bearish Bias)"
            else:
                sinyal_display = "HOLD"
        early_entry_warning = ""
    else:
        sinyal_display = sinyal
        # Early entry warning
        if "EARLY" in sinyal:
            early_entry_warning = "\n⚠️ EARLY ENTRY: Belum ada breakout/breakdown terkonfirmasi. Tunggu konfirmasi sebelum entry!"
        else:
            early_entry_warning = ""

# PRIORITAS UTAMA: Null/Undefined Fallback
    if "HOLD" in sinyal.upper() or entry == 0 or float(confidence) == 0.0 or "IGNORE" in sinyal.upper():
        entry_display = "-"
        sl_display = "-"
        tp_display = "-"
        risk_level = "LOW (Standby)" if "HOLD" in sinyal.upper() else "LOW (Avoid)"
        rr_ratio = "-"
        copy_trade_status = "NO TRADE"
        position_size_display = "-"
        risk_reward_display = "-"
        risk_pct_display = "-"
        kelly_display = "-"
    else:
        is_trade = "BUY" in sinyal or "SELL" in sinyal
        copy_trade_status = "OPEN" if is_trade else "NO TRADE"

    # Override for confidence-based status
    if confidence >= 60:
        copy_trade_status = "PENDING EXECUTION (Waiting Trigger)"
        position_size_display = f"{position_size:,.6f}" if position_size and position_size > 0 else "-"
        risk_reward_display = str(risk_metrics.get("risk_reward", 0))
        risk_pct_val = risk_metrics.get("risk_pct", 0)
        risk_pct_display = f"{risk_pct_val}%" if risk_pct_val > 0 else "-"
        kelly_val = risk_metrics.get("kelly_pct", 0)
        kelly_display = f"{kelly_val}%" if kelly_val > 0 else "-"
    elif confidence >= 40:
        copy_trade_status = "CONDITIONAL WATCHLIST"
        position_size_display = f"{position_size:,.6f}" if position_size and position_size > 0 else "-"
        risk_reward_display = str(risk_metrics.get("risk_reward", 0))
        risk_pct_val = risk_metrics.get("risk_pct", 0)
        risk_pct_display = f"{risk_pct_val}%" if risk_pct_val > 0 else "-"
        kelly_val = risk_metrics.get("kelly_pct", 0)
        kelly_display = f"{kelly_val}%" if kelly_val > 0 else "-"
    
    # Data Quality SELALU tampilkan (bukan hanya saat trading)
    data_score = data_quality.get("quality_score", 0)
    data_grade = data_quality.get("quality_grade", "N/A")
    if data_score > 0:
        data_quality_display = f"{data_grade} ({data_score}%)"
    else:
        data_quality_display = "N/A"
    
    # Jika ada warnings, tampilkan
    if data_quality.get("warnings"):
        warnings_text = " | ".join(data_quality.get("warnings", []))
        data_quality_display = f"{data_quality_display} | {warnings_text}"
    
    # Format berita dengan label
    berita_list = sentimen.get("berita_label", [])
    if berita_list:
        berita_str = "\n".join([f"  • [{b['label']}] {b['judul'][:60]}..." for b in berita_list[:3]])
    else:
        berita_str = "Tidak ada berita signifikan"
    
    # Advanced indicators display
    bb_pos = indikator.get("bb_position", "N/A") if indikator else "N/A"
    stoch = f"{indikator.get('stoch_k', 'N/A')}/{indikator.get('stoch_d', 'N/A')}" if indikator else "N/A"
    stoch_status = indikator.get('stoch_status', 'N/A') if indikator else "N/A"
    adx = indikator.get('adx', 0) if indikator else 0
    adx_status = indikator.get('adx_status', 'N/A') if indikator else "N/A"
    atr = indikator.get('atr', 0) if indikator else 0
    vwap = indikator.get('vwap', 0) if indikator else 0
    
    # Fibonacci levels
    fib_382 = indikator.get('fib_382', 0) if indikator else 0
    fib_500 = indikator.get('fib_500', 0) if indikator else 0
    fib_618 = indikator.get('fib_618', 0) if indikator else 0
    
    # Pivot Points
    pivot = indikator.get('pivot', 0) if indikator else 0
    r1 = indikator.get('r1', 0) if indikator else 0
    s1 = indikator.get('s1', 0) if indikator else 0
    
    # Validasi RR
    rr_valid_icon = "✓" if risk_metrics.get("is_valid_rr", False) else "✗"
    rr_final_display = f"{rr_ratio} {rr_valid_icon}" if rr_ratio != "-" else "-"

    # --- SAFETY NET: MENCEGAH UNBOUND LOCAL ERROR SAAT HOLD/NO TRADE ---
    if 'entry_display' not in locals(): entry_display = "-"
    if 'sl_display' not in locals(): sl_display = "-"
    if 'risk_pct_display' not in locals(): risk_pct_display = "-"
    if 'tp_display' not in locals(): tp_display = "-"
    if 'rr_ratio' not in locals(): rr_ratio = "-"
    if 'risk_reward_display' not in locals(): risk_reward_display = "-"
    if 'rr_valid_icon' not in locals(): rr_valid_icon = "✗"
    if 'kelly_display' not in locals(): kelly_display = "-"
    if 'risk_level' not in locals(): risk_level = "LOW (Standby)"
    if 'position_size_display' not in locals(): position_size_display = "-"
    # --------------------------------------------------------------------

    output = f"""📊 ANALISIS {symbol.upper()} | Data Quality: {data_quality_display}
💰 Harga: {harga_display}
🏪 Market: {market_condition} | ADX: {adx} ({adx_status})

============================================================
📈 INDIKATOR TEKNIKAL
============================================================
• RSI (14):     {indikator.get('rsi', 'N/A')} → {indikator.get('rsi_status', 'N/A')} | Skor: {skor_detail.get('rsi', 0)}
• MACD:         {indikator.get('macd_status', 'N/A')} | Histogram: {indikator.get('macd_hist_status', 'N/A')} | Skor: {skor_detail.get('macd', 0)}
• Stochastic:   {stoch} → {stoch_status}
• Bollinger:    {bb_pos}
• Trend:        {indikator.get('trend_status', 'N/A')} | MA20: {ma20_str} | MA50: {ma50_str} | MA200: {ma200_str} | Skor: {skor_detail.get('trend', 0)}
• Volume:       {indikator.get('volume_status', 'N/A')} (rasio: {indikator.get('volume_ratio', 'N/A')}x) | Skor: {skor_detail.get('volume', 0)}
• OBV Flow:     {indikator.get('obv_divergence', 'N/A')} | OBV: {indikator.get('obv', 0):,.0f}
• ATR:          {atr:.4f} | VWAP: {vwap:.4f}

============================================================
🎯 LEVEL & TARGET
============================================================
• S/R:          Support {s_display} | Resistance {r_display}
• Fibonacci:    38.2%: {fib_382:.4f} | 50%: {fib_500:.4f} | 61.8%: {fib_618:.4f}
• Pivot Points: PP: {pivot:.4f} | R1: {r1:.4f} | S1: {s1:.4f}

============================================================
📰 SENTIMEN BERITA
============================================================
{berita_str}
• Status: {sentimen.get('status', 'N/A')} | Skor: {skor_detail.get('sentimen', 0)}
• Dampak: {sentimen.get('dampak', 'N/A')}

============================================================
🎯 SIGNAL & SKOR
============================================================
Sinyal: {sinyal_display}
Skor Total: {'+' if float(total_skor) > 0 else ''}{total_skor} (RSI:{skor_detail.get('rsi', 0)} + MACD:{skor_detail.get('macd', 0)} + Trend:{skor_detail.get('trend', 0)} + Volume:{skor_detail.get('volume', 0)} + Sentimen:{skor_detail.get('sentimen', 0)})
Confidence: {confidence:.0f}%{no_trade_warning}{early_entry_warning}{weights_note}

============================================================
📊 RISK MANAGEMENT
============================================================
• Entry:        {entry_display}
• Stop Loss:    {sl_display} (Risk: {risk_pct_display})
• Take Profit:  {tp_display}
• Risk/Reward:  {rr_final_display}
• Kelly:        {kelly_display}
• Risk Level:   {risk_level}

============================================================
💼 POSITION SIZING
============================================================
• Modal:        Rp {modal:,.0f}
• Size:         {position_size_display}
• Status:       {copy_trade_status}

============================================================
🧠 ANALISIS
============================================================
{alasan}

============================================================
⚠️ DISCLAIMER: Bukan nasihat keuangan. Selalu lakukan riset sendiri dan kelola risiko dengan bijak."""
    
    return output

# ==============================
# ANALISIS AI V2 — MAIN ORCHESTRATOR
# ==============================
def analisis_ai_v2(symbol, jenis, data_harga, berita, indikator, modal=DEFAULT_MODAL):
    """
    Fungsi analisis utama — 10-Point System

    Returns: (formatted_output_string, analysis_data_dict)
    """
    # --- ULTIMATE FAILSAFE INITIALIZATION ---
    risk_pct_display = "-"
    ui_entry = "-"
    ui_sl = "-"
    ui_tp = "-"
    ui_rr = "-"
    ui_status = "STANDBY"
    # ----------------------------------------

    # --- GLOBAL PATCH UNTUK OBV & TIPE DATA ---
    try:
        obv_val = float(indikator.get("obv", 0))
        if indikator.get("obv_flow", "NONE") in ["NONE", ""]:
            indikator["obv_flow"] = "BULLISH (Accumulation)" if obv_val > 0 else "BEARISH (Distribution)"
    except Exception:
        pass
    # ------------------------------------------

    if not indikator:
        empty_data = {
            "sinyal": "HOLD", "entry": 0, "sl": 0, "tp": 0,
            "skor": 0, "confidence": 50, "risk_level": "LOW",
            "skor_detail": {}, "position_size": 0
        }
        return "⚠️ Data indikator tidak tersedia. Coba lagi nanti.", empty_data

    # Fallback Initialization
    execution_status = "NO TRADE"
    edge_clarity = "NO EDGE (Menunggu konfirmasi)"
    alasan = "Menunggu konfirmasi pergerakan harga lebih lanjut."
    next_trade_plan = "Standby. Evaluasi ulang setelah ada konfirmasi dari indikator volume dan tren."
    current_price = indikator.get("current_price", 0)
    plan_entry = current_price if current_price > 0 else 0
    plan_sl = plan_entry * 0.98
    plan_tp = plan_entry * 1.05
    
    # 1. Ambil harga saat ini
    if jenis == "Crypto":
        aset_key = list(data_harga.keys())[0]
        harga = data_harga[aset_key].get("usd", 0)
        harga_idr = data_harga[aset_key].get("idr", 0)
    else:
        harga = data_harga.get("harga", 0)
        harga_idr = None
    
    # 2. Adaptive Learning — ambil bobot dari histori
    trade_history = get_trade_history()
    weights = calculate_adaptive_weights(trade_history)
    
    # 3. Market condition (sudah dihitung di indikator)
    market_condition = indikator.get("market_condition", "UNKNOWN")

    # Determine trend_direction for contextual scoring
    current_price = indikator.get("current_price", 0)
    ma20 = indikator.get("ma20", 0)
    ma50 = indikator.get("ma50", 0)
    if current_price > ma20 and current_price > ma50:
        trend_direction = "BULLISH (Short-Medium)"
    elif current_price < ma20 and current_price < ma50:
        trend_direction = "BEARISH (Short-Medium)"
    else:
        trend_direction = "SIDEWAYS / TRANSISI"

    # 4. Analisis Berita
    sentimen = analyze_news(berita, symbol)

    # === VOLUME RATIO (SEBELUM SCORE) ===
    volume_ratio = indikator.get("volume_ratio", 1.0)

    # === VOLUME CONFIRMATION FOR SENTIMENT (SEBELUM SCORE) ===
    # Jika sentimen positif/negatif tapi Volume Ratio < 0.8x → downgrade ke NETRAL
    sentimen_skor = sentimen.get("skor", 0)
    if sentimen_skor != 0 and volume_ratio > 0:
        if volume_ratio < 0.8:
            sentimen["status"] = "NETRAL"
            sentimen["skor"] = 0
            sentimen["dampak"] = "News diabaikan pasar (Volume rendah)"

    # 5. Calculate Scores (deterministik)
    skor_detail, total_skor = calculate_score(indikator, sentimen, weights, market_condition, trend_direction)
    
    # === TENTUKAN TREND & VOLUME SEBELUM SINYAL ===
    is_trending_down = market_condition.startswith("TRENDING DOWN")
    is_trending_up = market_condition.startswith("TRENDING UP")
    vol_tinggi = indikator.get("volume_status") == "TINGGI"
    vol_rendah = indikator.get("volume_status") == "RENDAH"
    is_bearish = total_skor < 0
    is_bullish = total_skor > 0
    
    # === VOLUME RATIO FOR SIGNAL ENGINE ===
    volume_ratio = indikator.get("volume_ratio", 1.0)
    
    # === PRICE NEAR SUPPORT/RESISTANCE CHECK ===
    # Untuk Speculative Buy: harga dekat support (within 1.5%)
    # Untuk Speculative Sell: harga dekat resistance (within 1.5%)
    current_price = harga  # Gunakan variabel harga yang sudah ada
    support = indikator.get("support", 0)
    resistance = indikator.get("resistance", 0)
    price_near_support = False
    price_near_resistance = False
    
    if current_price > 0 and support > 0:
        # Harga dekat support jika within 1.5% dari support
        price_near_support = current_price <= support * 1.015
    
    if current_price > 0 and resistance > 0:
        # Harga dekat resistance jika within 1.5% dari resistance
        price_near_resistance = current_price >= resistance * 0.985
    
    # === KONFLIK INDIKATOR - JIKA KONFLIK: HOLD SAMPAI KONFIRMASI ===
    # Cek konflik antara trend dan momentum
    rsi_status = indikator.get("rsi_status", "NORMAL")
    macd_status = indikator.get("macd_status", "NORMAL")
    trend_status = indikator.get("trend_status", "NEUTRAL")
    
    # Deteksi konflik indikator
    has_conflict = False
    if ("BULLISH" in trend_status and "BEARISH" in macd_status) or ("BEARISH" in trend_status and "BULLISH" in macd_status):
        has_conflict = True
    if ("OVERSOLD" in rsi_status and "BEARISH" in macd_status) or ("OVERBOUGHT" in rsi_status and "BULLISH" in macd_status):
        has_conflict = True
    
    # === BREAKOUT/BREAKDOWN DETECTION ===
# Wajib untuk konfirmasi sinyal BUY/SELL
    breakout_status = detect_breakout_breakdown(indikator, market_condition)
    is_breakout = breakout_status == "breakout"
    is_breakdown = breakout_status == "breakdown"
    
    # === ENHANCED SIGNAL ENGINE ===
    # Speculative Buy (Buy on Weakness): Jika trend bearish + RSI oversold + harga dekat support
    # Strong Buy: Jika breakout + volume tinggi
    # Strong Sell: Jika breakdown + volume tinggi
    
    if has_conflict:
        sinyal = "HOLD"
        is_early_entry = False
    elif vol_rendah:
        sinyal = "HOLD"
        is_early_entry = False
    elif is_breakout and total_skor >= 1 and volume_ratio > 1.2:
        # Strong Buy: Breakout + Volume tinggi (1.2x+) + skor positif
        sinyal = "STRONG BUY (BREAKOUT CONFIRMED)"
        is_early_entry = False
    elif is_breakdown and total_skor <= -1 and volume_ratio > 1.2:
        # Strong Sell: Breakdown + Volume tinggi (1.2x+) + skor negatif
        sinyal = "STRONG SELL (BREAKDOWN CONFIRMED)"
        is_early_entry = False
    elif is_breakout and total_skor >= 1:
        # Regular Buy: Breakout + skor positif
        sinyal = "BUY (CONFIRMED)"
        is_early_entry = False
    elif is_breakdown and total_skor <= -1:
        # Regular Sell: Breakdown + skor negatif
        sinyal = "SELL (CONFIRMED)"
        is_early_entry = False
    elif is_trending_down and rsi_status == "OVERSOLD" and price_near_support:
        # Speculative Buy: Trend bearish + RSI oversold + harga dekat support
        # Ini adalah "buy on weakness" - accumulation zone
        sinyal = "SPECULATIVE BUY (Accumulation Zone)"
        is_early_entry = False
    elif is_trending_up and rsi_status == "OVERBOUGHT" and price_near_resistance:
        # Speculative Sell: Trend bullish + RSI overbought + harga dekat resistance
        # Ini adalah "sell on strength" - distribution zone
        sinyal = "SPECULATIVE SELL (Distribution Zone)"
        is_early_entry = False
    else:
        # TIDAK ADA KONFIRMASI → HOLD (Tidak ada EARLY)
        sinyal = "HOLD"
        is_early_entry = False
    
    # === SENTIMENT CONSISTENCY ===
    sentimen_status = sentimen.get("status", "NETRAL")
    sentimen_skor = sentimen.get("skor", 0)
    if sentimen_status == "NETRAL" and sentimen_skor != 0:
        # Jika NETRAL tapi skor ≠ 0 → ubah ke NETRAL dengan skor 0
        sentimen["skor"] = 0
        sentimen["dampak"] = "Tidak ada sentimen signifikan"
    elif sentimen_status != "NETRAL" and sentimen_skor == 0:
        # Jika tidak NETRAL tapi skor = 0 → ubah ke NETRAL
        sentimen["status"] = "NETRAL"
        sentimen["dampak"] = "Tidak ada sentimen signifikan"
    
    # Cek apakah ada berita untuk confidence adjustment
    has_news = berita and len(berita) > 0
    
    # Cek trend dan struktur S/R
    has_clear_trend = is_trending_up or is_trending_down
    has_valid_sr = indikator.get("support", 0) > 0 and indikator.get("resistance", 0) > 0
    
    # Cek konflik untuk confidence scoring (sudah dihitung di sinyal logic)
    # has_conflict sudah ada dari sinyal logic
    
# 7. CONFIDENCE SCORE (DYNAMIC)
    # Determine if trend_direction aligned with signal
    trend_bullish = "BULLISH" in trend_direction
    trend_bearish = "BEARISH" in trend_direction
    signal_buy = "BUY" in sinyal.upper()
    signal_sell = "SELL" in sinyal.upper()
    signal_hold = "HOLD" in sinyal.upper()

    aligned = False
    if trend_bullish and (signal_buy or signal_hold):
        aligned = True
    elif trend_bearish and (signal_sell or signal_hold):
        aligned = True

    # Strategy regime
    strategy_regime = "RANGE TRADING" if "SIDEWAYS" in market_condition else "TREND TRADING"

    # Check for conflicts (MACD vs OBV, or Signal vs Trend)
    macd_status = indikator.get("macd_status", "")
    obv_divergence = indikator.get("obv_divergence", "")
    macd_bullish = "BULLISH" in macd_status
    macd_bearish = "BEARISH" in macd_status
    obv_bullish = "BULLISH" in obv_divergence
    obv_bearish = "BEARISH" in obv_divergence

    macd_obv_conflict = (macd_bullish and obv_bearish) or (macd_bearish and obv_bullish)
    signal_trend_conflict = not aligned and (signal_buy or signal_sell)

    many_conflicts = macd_obv_conflict or signal_trend_conflict

    if aligned:
        confidence = 70
    elif strategy_regime == "RANGE TRADING" and not many_conflicts:
        confidence = 60
    elif many_conflicts:
        confidence = 35
    else:
        confidence = 50
    
    # 8. No-Trade Zone Check
    no_trade = detect_no_trade_zone(indikator, skor_detail)
    if no_trade:
        sinyal = "HOLD"
    
    # 9. Entry, SL, TP
    entry, sl, tp = calculate_entry_sl_tp(harga, sinyal, indikator)
    
    # 10. Risk & Reward
    is_buy = "BUY" in sinyal
    is_sell = "SELL" in sinyal
    
    if is_buy and entry > 0:
        risk_pct = abs(entry - sl) / entry * 100
        reward_pct = abs(tp - entry) / entry * 100
    elif is_sell and entry > 0:
        risk_pct = abs(sl - entry) / entry * 100
        reward_pct = abs(entry - tp) / entry * 100
    else:
        risk_pct = 0
        reward_pct = 0
    
    # OPTIMASI TP: Jika RR < 1:1.5, sesuaikan TP agar minimal 1:1.5
    if risk_pct > 0 and reward_pct > 0:
        current_rr = reward_pct / risk_pct
        if current_rr < 1.5:
            # Sesuaikan TP untuk RR minimal 1:1.5
            if is_buy:
                tp = entry + (entry - sl) * 1.5
            elif is_sell:
                tp = entry - (sl - entry) * 1.5
            reward_pct = abs(tp - entry) / entry * 100 if is_buy else abs(entry - tp) / entry * 100
    
    # RISK / REWARD WAJIB VALID
    if risk_pct > 0 and reward_pct > 0:
        final_rr = reward_pct / risk_pct
        if final_rr >= 1.5:
            rr_ratio = f"1:{final_rr:.1f}"
        else:
            rr_ratio = "N/A"
    else:
        rr_ratio = "N/A"
    
    # === VALIDASI OTOMATIS ===
    # Pastikan arah BUY/SELL benar
    if is_buy:
        # BUY: TP > Entry > SL
        if not (tp > entry and entry > sl):
            # Perbaiki otomatis
            sl = entry * 0.98  # 2% di bawah entry
            tp = entry * 1.06  # 6% di atas entry (RR 1:3)
    elif is_sell:
        # SELL: SL > Entry > TP
        if not (sl > entry and entry > tp):
            # Perbaiki otomatis
            sl = entry * 1.02  # 2% di atas entry
            tp = entry * 0.94  # 6% di bawah entry (RR 1:3)
    
    # Risk Level - Dynamic Assessment
    current_price = indikator.get("current_price", 0)
    resistance = indikator.get("resistance", 0)
    support = indikator.get("support", 0)
    ma20 = indikator.get("ma20", 0)
    vol_status = indikator.get("volume_status", "NORMAL")
    adx = indikator.get("adx", 0)
    is_bullish_trend = market_condition.startswith("TRENDING UP")
    is_sideways = "SIDEWAYS" in market_condition

    # HIGH RISK: Volume tinggi or harga <2% dari Resistance
    near_resistance = resistance > 0 and abs(current_price - resistance) / resistance <= 0.02
    if vol_status == "TINGGI" or near_resistance:
        risk_level = "HIGH"
    # MEDIUM RISK: Trend kuat (ADX > 25) or Sideways normal
    elif adx > 25 or is_sideways:
        risk_level = "MEDIUM"
    # LOW RISK: Setup Pullback Sempurna (Harga dekat MA20/Support di Uptrend)
    elif is_bullish_trend and (support > 0 and abs(current_price - support) / support <= 0.02 or ma20 > 0 and abs(current_price - ma20) / ma20 <= 0.02):
        risk_level = "LOW"
    else:
        risk_level = "MEDIUM"  # Default
    
    # Position Sizing
    # Tentukan currency berdasarkan jenis aset
    if jenis == "Crypto":
        curr = "USD"
    elif symbol.endswith(".JK"):
        curr = "IDR"
    else:
        curr = "USD"
    
    position_size = calculate_position_size(modal, entry, sl, sinyal, market_condition, curr)
    
    # POSITION SIZE VALIDATION: Jika size < 1 → NO TRADE
    if position_size < 1 and sinyal != "HOLD":
        sinyal = "HOLD"
        entry = 0
        sl = 0
        tp = 0
        position_size = 0
    
# Calculate professional risk metrics
    risk_metrics = calculate_risk_metrics(harga, entry, sl, tp, indikator, sinyal, modal, curr)
    
    # Calculate data quality
    data_quality = calculate_data_quality(indikator, berita)
    
    # ============================================================
    # VALIDASI OTOMATIS SEBELUM OUTPUT (KETAT)
    # ============================================================
    
    # 1. BIAS LOCK - Pastikan Bias sesuai Trend + ADX
    adx = indikator.get("adx", 0)
    trend_status = indikator.get("trend_status", "NEUTRAL")
    current_price = indikator.get("current_price", 0)
    ma50 = indikator.get("ma50", 0)
    ma200 = indikator.get("ma200", 0)
    
    # Tentukan trend direction yang dominan
    trend_is_bearish = (trend_status == "BEARISH") or (current_price < ma50 and ma50 < ma200 and ma200 > 0)
    trend_is_bullish = (trend_status == "BULLISH") or (current_price > ma50 and ma50 > ma200 and ma200 > 0)
    trend_is_strong = adx >= 25  # ADX >= 25 = trend kuat
    
    # Jika trend kuat (ADX >= 25) → bias WAJIB mengikuti trend
    if trend_is_strong:
        if trend_is_bearish and is_buy and not is_breakout:
            sinyal = "HOLD (Bearish Bias)"
            is_early_entry = False
            print(f"Validasi: Trend bearish kuat (ADX={adx}) → HOLD (Bearish Bias)")
        elif trend_is_bullish and is_sell and not is_breakdown:
            sinyal = "HOLD (Bullish Bias)"
            is_early_entry = False
            print(f"Validasi: Trend bullish kuat (ADX={adx}) → HOLD (Bullish Bias)")
    
    # 2. SCORE INTEGRITY - Validasi penjumlahan skor
    expected_total = round(skor_detail.get("rsi", 0) + skor_detail.get("macd", 0) + skor_detail.get("trend", 0) + skor_detail.get("volume", 0) + skor_detail.get("sentimen", 0), 1)
    if abs(expected_total - total_skor) > 0.1:
        print(f"Warning: Skor tidak konsisten. Expected: {expected_total}, Actual: {total_skor}")
        total_skor = expected_total
    
    # 3. DATA QUALITY CHECK - Deteksi anomali dan indikator identik
    ma20 = indikator.get("ma20", 0)
    data_anomaly = False
    
    # Check MA20 ≈ MA50
    if ma20 and ma50 and ma50 != 0:
        ma_diff_pct = abs(ma20 - ma50) / ma50 * 100
        if ma_diff_pct < 0.5:
            data_anomaly = True
    
    # Check indikator identik (RSI = Stochastic, dll)
    rsi = indikator.get("rsi", 0)
    stoch_k = indikator.get("stoch_k", 0)
    if rsi > 0 and stoch_k > 0 and abs(rsi - stoch_k) < 1:
        data_anomaly = True
    
    # Set quality warning dan turunkan confidence
    if data_anomaly:
        current_grade = data_quality.get("quality_grade", "C")
        data_quality["quality_grade"] = f"WARNING ({current_grade})"
        data_quality["warnings"] = data_quality.get("warnings", []) + ["Data tidak normal atau flat"]
        confidence = max(35, confidence - 10)
        print(f"Validasi: Data anomaly detected → Confidence diturunkan")
    
    # 4. SIGNAL CONSISTENCY - Skor vs Sinyal
    # Skor -1 sampai +1 → HOLD
    # Skor ≥ +4 → BUY (hanya jika breakout)
    # Skor ≤ -3 → SELL (hanya jika breakdown)
    if -1 <= total_skor <= 1:
        sinyal = "HOLD"
        is_early_entry = False
        print(f"Validasi: Skor lemah ({total_skor}) → HOLD")
    elif total_skor >= 4 and not is_breakout:
        # Skor kuat ≥4 tapi belum breakout → HOLD
        sinyal = "HOLD"
        is_early_entry = False
        print(f"Validasi: Skor kuat ({total_skor}) tapi belum breakout → HOLD")
    elif total_skor <= -3 and not is_breakdown:
        # Skor kuat negative tapi belum breakdown → HOLD
        sinyal = "HOLD"
        is_early_entry = False
        print(f"Validasi: Skor negatif ({total_skor}) tapi belum breakdown → HOLD")
    
    # 4b. TREND CONSISTENCY VALIDATION - Harga vs MA
    # Jika harga < MA50 dan MA200 → trend TIDAK BOLEH bullish
    current_price = indikator.get("current_price", 0)
    ma50 = indikator.get("ma50", 0)
    ma200 = indikator.get("ma200", 0)
    trend_status = indikator.get("trend_status", "NEUTRAL")
    
    if current_price > 0 and ma50 > 0 and ma200 > 0:
        # Cek posisi harga vs MA
        price_below_ma50 = current_price < ma50
        price_below_ma200 = current_price < ma200
        ma20 = indikator.get("ma20", 0)
        
        # Jika harga di bawah MA50 dan MA200 → trend harus bearish
        if price_below_ma50 and price_below_ma200:
            if "BULLISH" in trend_status:
                trend_status = "BEARISH"
                indikator["trend_status"] = "BEARISH"
                print("Validasi: Harga di bawah MA50 & MA200 → Override trend ke BEARISH")
        
        # Cek MA structure
        if ma20 > 0 and ma50 > 0 and ma200 > 0:
            if ma20 < ma50 < ma200:
                # Strong bearish structure
                trend_status = "BEARISH"
                indikator["trend_status"] = "BEARISH"
                print("Validasi: MA20 < MA50 < MA200 → Strong BEARISH")
    
    # 5. RISK/REWARD VALIDATION - Jika tidak ada entry
    if sinyal == "HOLD" or entry == 0 or entry is None:
        risk_metrics["rr_ratio"] = "N/A"
        risk_metrics["is_valid_rr"] = False
        entry = 0
        sl = 0
        tp = 0
        position_size = 0
    
    # 6. NO CONTRADICTION - Pastikan sinyal tidak berkontradiksi dengan trend
    # Jika trend bearish kuat tapi sinyal bullish → override
    if trend_is_strong:
        if trend_is_bearish and "BUY" in sinyal and "CONFIRMED" not in sinyal:
            sinyal = "HOLD (Bearish Bias)"
            is_early_entry = False
            print("Validasi: Kontradiksi trend bearish + sinyal bullish → HOLD")
        elif trend_is_bullish and "SELL" in sinyal and "CONFIRMED" not in sinyal:
            sinyal = "HOLD (Bullish Bias)"
            is_early_entry = False
            print("Validasi: Kontradiksi trend bullish + sinyal bearish → HOLD")
    
    # 7. Re-capture signals after validation
    is_buy = "BUY" in sinyal
    is_sell = "SELL" in sinyal
    
    # 8. FORCE BIAS ON HOLD - Jika HOLD wajib ada bias
    if sinyal == "HOLD" and "(" not in sinyal:
        if trend_is_bearish:
            sinyal = "HOLD (Bearish Bias)"
        elif trend_is_bullish:
            sinyal = "HOLD (Bullish Bias)"
        else:
            # Weak trend - gunakan skor untuk tentukan bias
            if total_skor > 0:
                sinyal = "HOLD (Bullish Bias)"
            elif total_skor < 0:
                sinyal = "HOLD (Bearish Bias)"
    
    # 9. FINAL SANITY CHECK
    # Pastikan tidak ada konflik antara skor dan sinyal
    if is_buy and total_skor < 0:
        print(f"Warning: Sinyal BUY tapi skor negatif ({total_skor})")
        sinyal = "HOLD"
        is_buy = False
    if is_sell and total_skor > 0:
        print(f"Warning: Sinyal SELL tapi skor positif ({total_skor})")
        sinyal = "HOLD"
        is_sell = False
    

    
    # Final contradiction check - perbaiki alasan jika berkontradiksi
    # Jika alasan menyebut bullish tapi trend bearish kuat → perbaiki
    if trend_is_strong and trend_is_bearish and "bullish" in alasan.lower() and "hold" not in sinyal.lower():
        # Override reason untuk konsistensi
        alasan = f"HOLD karena trend bearish kuat (ADX={adx}). Harga di bawah MA50 & MA200 menunjukkan dominasi seller. Tunggu pullback ke resistance {indikator.get('resistance', 0):.2f} dengan volume tinggi untuk konfirmasi."
    elif trend_is_strong and trend_is_bullish and "bearish" in alasan.lower() and "hold" not in sinyal.lower():
        alasan = f"HOLD karena trend bullish kuat (ADX={adx}). Harga di atas MA50 & MA200 menunjukkan dominasi buyer. Tunggu pullback ke support {indikator.get('support', 0):.2f} dengan volume tinggi untuk konfirmasi."

    # === URGENT ARCHITECTURE FIX: DECOUPLING MARKET VS ENTRY QUALITY ===

    # Evaluasi Market Quality dari Skor Total dan Trend
    market_quality = "BURUK"
    if total_skor > 3 and "BULLISH" in trend_direction:
        market_quality = "BAGUS"

    # Hitung R:R secara real-time
    if rr_ratio and " : " in rr_ratio:
        rr_numeric = float(rr_ratio.split(" : ")[1])
    else:
        rr_numeric = 0

    entry_quality = "BAGUS" if rr_numeric >= 1.5 else "BURUK"

    # Sinkronisasi Mutlak
    if entry_quality == "BURUK" or rr_numeric < 1.5:
        execution_status = "NO TRADE"
        sinyal = "WAIT (No Valid Setup)"
        risk_level = "NONE (Di Luar Market)"
        if total_skor >= 3:
            confidence = 45  # Weak Edge / Tren Kuat
            edge_clarity = "WEAK EDGE (Arah market jelas, tapi zona Entry berisiko tinggi / R:R buruk)."
        else:
            confidence = 35  # No Edge / Choppy
            edge_clarity = "NO EDGE (Market tidak jelas dan tidak ada momentum valid)."
        rr_display = "N/A (Tidak ada setup valid)"
        execution_setup = "NO TRADE. Abaikan spekulasi."
    elif entry_quality == "BAGUS" and (("BUY" in sinyal.upper() and "BULLISH" in trend_direction) or ("SELL" in sinyal.upper() and "BEARISH" in trend_direction)):
        execution_status = "READY TO EXECUTE"
        vol_status = indikator.get("volume_status", "NORMAL")
        adx = indikator.get("adx", 0)
        if vol_status == "TINGGI" or adx > 30:
            risk_level = "HIGH"
        else:
            risk_level = "MEDIUM"
        confidence = 75
    else:
        execution_status = "HOLD"

    # === THE "NEXT TRADE PLAN" ENGINE (PROBABILITY & OBV FILTER) ===
    adx = indikator.get("adx", 0)
    support = indikator.get("support", 0)
    resistance = indikator.get("resistance", 0)
    current_price = indikator.get("current_price", 0)
    ma20 = indikator.get("ma20", 0)
    ma50 = indikator.get("ma50", 0)
    vwap = indikator.get("vwap", 0)
    obv_flow = indikator.get("obv_divergence", "NONE")
    # Patch OBV Otomatis untuk mencegah "NONE"
    obv_val = float(indikator.get("obv", 0))
    if obv_flow == "NONE" or obv_flow == "":
        obv_flow = "BULLISH (Accumulation)" if obv_val > 0 else "BEARISH (Distribution)"

    is_bearish_macro = "TRENDING DOWN" in market_condition or (ma50 > 0 and current_price < ma50)
    is_bullish_macro = "TRENDING UP" in market_condition or (ma50 > 0 and current_price > ma50)

    setup_type = ""
    plan_entry = 0
    plan_sl = 0
    plan_tp = 0
    trigger = ""
    invalidasi = ""

    # 1. Regime Detection & Strict Triggers
    if adx >= 20:
        if is_bearish_macro:
            setup_type = "SELL ON RALLY"
            dynamic_res = ma20 if current_price < ma20 else vwap
            plan_entry = dynamic_res if dynamic_res > current_price else resistance
            plan_sl = plan_entry * 1.03
            plan_tp = support
            trigger = "Bearish Engulfing/Pin Bar + Vol > MA20"
            invalidasi = f"Close di atas {plan_sl:.2f}"
        elif is_bullish_macro:
            setup_type = "BUY ON DIP"
            dynamic_sup = ma20 if current_price > ma20 else vwap
            plan_entry = dynamic_sup if dynamic_sup > support else support
            plan_sl = plan_entry * 0.97
            plan_tp = resistance
            trigger = "Bullish Engulfing/Pin Bar + Vol > MA20"
            invalidasi = f"Close di bawah {plan_sl:.2f}"
    else:
        # SMART REACTIVE RANGE TRADING (Sideways/Choppy)
        jarak_ke_support = abs(current_price - support)
        jarak_ke_res = abs(resistance - current_price)

        # Deteksi Tekanan Tersembunyi (Hidden Pressure)
        if is_bearish_macro and "Distribution" in obv_flow:
            setup_type = "REACTIVE RANGE SELL (Searah Tren & OBV)"
            plan_entry = resistance
            plan_sl = resistance * 1.03
            plan_tp = support
            trigger = "Bearish Rejection Valid + Vol > MA20"
            invalidasi = f"Close kuat di atas {plan_sl:.2f}"
            skenario_alternatif = f"ALT: Counter-Trend BUY di {support:.2f} SANGAT BERISIKO karena tekanan Distribusi OBV. Rentan Breakdown."
        elif is_bullish_macro and "Accumulation" in obv_flow:
            setup_type = "REACTIVE RANGE BUY (Searah Tren & OBV)"
            plan_entry = support
            plan_sl = support * 0.97
            plan_tp = resistance
            trigger = "Bullish Rejection Valid + Vol > MA20"
            invalidasi = f"Close kuat di bawah {plan_sl:.2f}"
            skenario_alternatif = f"ALT: Counter-Trend SELL di {resistance:.2f} SANGAT BERISIKO karena tekanan Akumulasi OBV. Rentan Breakout."
        else:
            # Sideways Murni (Netral), pilih yang terdekat
            if jarak_ke_support <= jarak_ke_res:
                setup_type = "REACTIVE RANGE BUY (Support Bounce)"
                plan_entry = support
                plan_sl = support * 0.97
                plan_tp = resistance
                trigger = "Bullish Rejection Valid di Support + Vol > MA20"
                invalidasi = f"Close kuat di bawah {plan_sl:.2f}"
                skenario_alternatif = f"ALT: Jika Breakdown Support, beralih ke SELL."
            else:
                setup_type = "REACTIVE RANGE SELL (Resistance Rejection)"
                plan_entry = resistance
                plan_sl = resistance * 1.03
                plan_tp = support
                trigger = "Bearish Rejection Valid di Resistance + Vol > MA20"
                invalidasi = f"Close kuat di atas {plan_sl:.2f}"
                skenario_alternatif = f"ALT: Jika Breakout Resistance, beralih ke BUY."

    # 2. Probability & R:R Calculation
    potensi_rr = 0
    if plan_entry > 0 and plan_sl > 0 and plan_entry != plan_sl:
        risk = abs(plan_entry - plan_sl)
        reward = abs(plan_tp - plan_entry)
        potensi_rr = reward / risk if risk > 0 else 0

    jarak_entry_pct = abs(current_price - plan_entry) / current_price * 100 if current_price > 0 else 0

    skenario_alternatif = ""
    if adx < 20: # Hanya di pasar sideways
        if "SELL" in setup_type:
            skenario_alternatif = f"ALT Skenario: JIKA Breakout Resistance, beralih ke BUY di area {resistance:.2f}."
        else:
            skenario_alternatif = f"ALT Skenario: JIKA Breakdown Support, beralih ke SELL di area {support:.2f}."

    # 3. Decision Driver & Confidence Matrix
    if execution_status == "NO TRADE":
        if potensi_rr < 1.5:
            edge_clarity = "NO ENTRY YET (Proyeksi R:R Buruk)"
            next_trade_plan = f"TAHAN EKSEKUSI. R:R saat ini (1:{potensi_rr:.1f}) di bawah standar institusi. Menunggu harga terkoreksi ke area yang lebih aman."
            confidence = 0
        else:
            calc_conf = 50 + int(potensi_rr * 5) - int(jarak_entry_pct * 2)

            # CONFIDENCE PENALTY: Diskon keyakinan jika melawan arus atau jarak jauh
            if "Counter-Trend" in skenario_alternatif or ("Bearish" in market_condition and "BUY" in setup_type):
                calc_conf -= 20 # Penalti berat melawan tren
            if jarak_entry_pct > 4.0:
                calc_conf -= 15 # Penalti jarak

            confidence = max(0, min(80, calc_conf))

            if confidence < 40:
                edge_clarity = "NEUTRAL WATCH (Risiko Tinggi/Harga di Tengah)"
                next_trade_plan = f"TAHAN EKSEKUSI. Harga mengambang di 'No Man's Land' atau berisiko tinggi. STRATEGI: {setup_type} | PANTAU: {plan_entry:.2f} | R:R 1:{potensi_rr:.1f}. {skenario_alternatif}"
            elif confidence <= 60:
                edge_clarity = "ROADMAP (Watchlist Menengah)"
                next_trade_plan = f"STRATEGI: {setup_type}. PANTAU: {plan_entry:.2f} (Jarak: {jarak_entry_pct:.1f}%). TRIGGER: {trigger}. TP: {plan_tp:.2f} | SL: {plan_sl:.2f} | R:R 1:{potensi_rr:.1f}. {skenario_alternatif}"
            else:
                edge_clarity = "READY (High Probability Setup)"
                next_trade_plan = f"STRATEGI: {setup_type}. PANTAU KETAT: {plan_entry:.2f} (Jarak ideal: {jarak_entry_pct:.1f}%). TRIGGER: {trigger}. TP: {plan_tp:.2f} | SL: {plan_sl:.2f} | R:R 1:{potensi_rr:.1f}. BATAL JIKA: {invalidasi}. {skenario_alternatif}"

    # --- GLOBAL UI SYNC UNTUK RISK MANAGEMENT & STATUS ---
    if confidence >= 60:
        ui_entry = f"{plan_entry:.2f}"
        ui_sl = f"{plan_sl:.2f}"
        ui_tp = f"{plan_tp:.2f}"
        ui_rr = f"1:{potensi_rr:.1f}"
        ui_status = "PENDING EXECUTION (Waiting Trigger)"
    elif confidence >= 40:
        ui_entry = f"{plan_entry:.2f}"
        ui_sl = f"{plan_sl:.2f}"
        ui_tp = f"{plan_tp:.2f}"
        ui_rr = f"1:{potensi_rr:.1f}"
        ui_status = "CONDITIONAL WATCHLIST"
    else:
        # Keep defaults
        pass

    # Convert ui_* to actual variables for format_analysis_output
    entry = float(ui_entry) if ui_entry != "-" else 0
    sl = float(ui_sl) if ui_sl != "-" else 0
    tp = float(ui_tp) if ui_tp != "-" else 0
    rr_ratio = ui_rr
    copy_trade_status = ui_status

    # CATATAN UNTUK AI IDE:
    # Pastikan variabel entry, sl, tp, rr_ratio, dan copy_trade_status
    # DIMASUKKAN ke dalam f-string pembentuk tabel laporan (Risk Management & Position Sizing).
    # -----------------------------------------------------

    # AI Reasoning (LLM call — hanya untuk penjelasan)
    # Gunakan ui_rr karena rr_display is undefined
    alasan = get_ai_reasoning(
        symbol, indikator, sentimen, skor_detail, total_skor,
        sinyal, market_condition, confidence, ui_rr,
        risk_level, execution_status, edge_clarity, next_trade_plan
    )

    # Format Output
    output = format_analysis_output(
        symbol=symbol,
        harga=harga,
        harga_idr=harga_idr,
        indikator=indikator,
        sentimen=sentimen,
        sinyal=sinyal,
        total_skor=total_skor,
        confidence=confidence,
        entry=entry,
        sl=sl,
        tp=tp,
        risk_level=risk_level,
        rr_ratio=rr_ratio,
        modal=modal,
        position_size=position_size,
        alasan=alasan,
        market_condition=market_condition,
        skor_detail=skor_detail,
        weights=weights,
        jenis=jenis,
        no_trade=no_trade,
        is_early_entry=is_early_entry,
        risk_metrics=risk_metrics,
        data_quality=data_quality
    )
    
    # ============================================================
    # SAFETY NET: STRICT VARIABLE BINDING & NULL FALLBACK
    # ============================================================
    if confidence == 0 or "IGNORE" in sinyal.upper() or "HOLD" in sinyal.upper():
        entry = 0
        sl = 0
        tp = 0
        rr_ratio = "-"
        position_size = 0
        risk_metrics["risk_reward"] = 0
        risk_metrics["kelly_pct"] = 0
        risk_metrics["rr_ratio"] = "-"
        risk_level = risk_level if risk_level else "LOW (Avoid)"
    
    analysis_data = {
        "sinyal": sinyal,
        "entry": entry,
        "sl": sl,
        "tp": tp,
        "skor": total_skor,
        "confidence": confidence,
        "risk_level": risk_level,
        "skor_detail": skor_detail,
        "position_size": position_size,
        "market_condition": market_condition,
        "no_trade": no_trade,
        "is_early_entry": is_early_entry,
        "breakout_status": breakout_status
    }
    
    return output, analysis_data

# ==============================
# LEGACY ANALYSIS (backward compat)
# ==============================
def analisis_ai(pertanyaan, data_harga, berita, indikator):
    """Legacy function — tetap ada untuk backward compatibility"""
    prompt = f"""
    Kamu adalah asisten investasi profesional yang membantu investor Indonesia.
    
    Data harga saat ini:
    {data_harga}
    
    Indikator Teknikal:
    {indikator}
    
    Berita terkini:
    {berita}
    
    Pertanyaan pengguna: {pertanyaan}

    Berikan analisis yang:
    1. Jelas dan mudah dipahami pemula
    2. Jelaskan kondisi RSI dan MACD dengan bahasa sederhana
    3. Hubungkan berita terkini dengan indikator teknikal
    4. Sertakan sinyal: BELI / JUAL / HOLD
    5. Sentimen pasar: POSITIF / NEGATIF / NETRAL
    6. Berikan rekomendasi trading:
       - Entry Price: harga ideal untuk masuk posisi
       - Stop Loss: batas harga keluar jika salah (sekitar 3-5% dari entry)
       - Take Profit: target harga ambil untung (sekitar 5-10% dari entry)
       - Risk Level: LOW / MEDIUM / HIGH
       - Risk/Reward Ratio: perbandingan risiko vs potensi untung
    7. Jelaskan alasannya secara singkat
    8. Ingatkan bahwa investasi selalu ada risikonya
    
    Jawab dalam Bahasa Indonesia.
    """
    
    response = client.chat.completions.create(
        model="meta-llama/llama-3.3-70b-instruct",
        max_tokens=1500,
        messages=[{"role": "user", "content": prompt}]
    )
    return response.choices[0].message.content

def parse_trading_info(hasil_analisis, harga_sekarang):
    """Ekstrak Entry, SL, TP dari teks analisis AI (legacy)"""
    entry = harga_sekarang
    sl = harga_sekarang * 0.95
    tp = harga_sekarang * 1.10
    sinyal = "HOLD"
    
    # Deteksi sinyal
    if "BELI" in hasil_analisis.upper():
        sinyal = "BELI"
    elif "JUAL" in hasil_analisis.upper():
        sinyal = "JUAL"
    
    # Cari angka setelah "Entry Price"
    entry_match = re.search(r'Entry Price[:\s]+Rp?\s?([\d.,]+)', hasil_analisis)
    sl_match = re.search(r'Stop Loss[:\s]+Rp?\s?([\d.,]+)', hasil_analisis)
    tp_match = re.search(r'Take Profit[:\s]+Rp?\s?([\d.,]+)', hasil_analisis)
    
    if entry_match:
        try:
            entry = float(entry_match.group(1).replace(".", "").replace(",", ""))
        except:
            pass
    if sl_match:
        try:
            sl = float(sl_match.group(1).replace(".", "").replace(",", ""))
        except:
            pass
    if tp_match:
        try:
            tp = float(tp_match.group(1).replace(".", "").replace(",", ""))
        except:
            pass
    
    return sinyal, entry, sl, tp

# ==============================
# FUNGSI GOOGLE SHEETS — SIMPAN
# ==============================
def simpan_ke_sheets(jenis, nama_aset, harga_data, analisis, indikator):
    try:
        gc = get_sheets_client()
        if not gc:
            return

        sheet_id = os.getenv("GOOGLE_SHEETS_ID")
        spreadsheet = gc.open_by_key(sheet_id)

        # Cek sheet Histori, kalau tidak ada buat baru
        try:
            ws = spreadsheet.worksheet("Histori")
        except:
            ws = spreadsheet.add_worksheet("Histori", 1000, 10)
            ws.append_row([
                "Tanggal & Waktu", "Jenis", "Nama Aset",
                "Harga", "RSI", "MACD", "Sinyal AI", "Analisis"
            ])

        # Ambil data harga
        try:
            if jenis == "Crypto":
                aset_key = list(harga_data.keys())[0]
                harga = f"IDR {harga_data[aset_key].get('idr', '-'):,} | USD {harga_data[aset_key].get('usd', '-'):,}"
            else:
                harga = f"{harga_data.get('mata_uang', '')} {harga_data.get('harga', '-')}"
        except:
            harga = "-"

        # Ambil indikator
        rsi = indikator.get("rsi", "-") if indikator else "-"
        macd = indikator.get("macd_status", "-") if indikator else "-"

        # Tentukan sinyal dari teks analisis
        sinyal = "HOLD"
        analisis_upper = analisis.upper() if isinstance(analisis, str) else ""
        if "BELI" in analisis_upper:
            sinyal = "BELI"
        elif "JUAL" in analisis_upper:
            sinyal = "JUAL"

        # Tambah baris baru
        ws.append_row([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            jenis,
            nama_aset.upper(),
            harga,
            str(rsi),
            macd,
            sinyal,
            analisis[:500] if isinstance(analisis, str) else str(analisis)[:500]
        ])
        print(f"✅ Analisis tersimpan ke Google Sheets!")

    except Exception as e:
        print(f"⚠️ Gagal simpan ke Sheets: {e}")

# ==============================
# FUNGSI PERFORMANCE TRACKING
# ==============================
def catat_sinyal(jenis, nama_aset, harga_entry, sinyal, stop_loss, take_profit, skor_detail=None):
    try:
        gc = get_sheets_client()
        if not gc:
            return

        sheet_id = os.getenv("GOOGLE_SHEETS_ID")
        spreadsheet = gc.open_by_key(sheet_id)

        # Header termasuk Skor Detail untuk adaptive learning
        headers = [
            "Tanggal", "Jenis", "Aset", "Sinyal",
            "Harga Entry", "Stop Loss", "Take Profit",
            "Harga Penutupan", "Hasil", "Profit/Loss %",
            "Status", "Skor Detail"
        ]
        try:
            ws = spreadsheet.worksheet("Performance")
            first_row = ws.row_values(1)
            if not first_row or first_row[0] != "Tanggal":
                ws.insert_row(headers, 1)
                print("✅ Header Performance ditambahkan!")
        except:
            ws = spreadsheet.add_worksheet("Performance", 1000, 15)
            ws.append_row(headers)

        ws.append_row([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            jenis,
            nama_aset.upper(),
            sinyal,
            harga_entry,
            stop_loss,
            take_profit,
            "-",      # Harga penutupan (diisi nanti)
            "-",      # Hasil (WIN/LOSS)
            "-",      # Profit/Loss %
            "OPEN",   # Status
            json.dumps(skor_detail) if skor_detail else ""
        ])
        print(f"✅ Sinyal dicatat ke Performance Tracker!")

    except Exception as e:
        print(f"⚠️ Gagal catat sinyal: {e}")

def hitung_performa():
    try:
        gc = get_sheets_client()
        if not gc:
            return None

        sheet_id = os.getenv("GOOGLE_SHEETS_ID")
        spreadsheet = gc.open_by_key(sheet_id)
        ws = spreadsheet.worksheet("Performance")
        
        # Gunakan get_all_values agar lebih robust
        all_values = ws.get_all_values()
        
        if len(all_values) < 2:
            return None
        
        # Cari header row
        header_idx = 0
        for i, row in enumerate(all_values):
            if "Tanggal" in row or "Status" in row:
                header_idx = i
                break
        
        headers = all_values[header_idx]
        data_rows = all_values[header_idx + 1:]
        
        if not data_rows:
            return None
        
        # Convert ke dict
        data = []
        for row in data_rows:
            if len(row) >= len(headers):
                d = dict(zip(headers, row))
                data.append(d)
            elif len(row) > 0 and row[0]:  # Ada data tapi kolom kurang
                padded = row + [""] * (len(headers) - len(row))
                d = dict(zip(headers, padded))
                data.append(d)
        
        if not data:
            return None

        total = len(data)
        closed = [d for d in data if d.get("Status", "") == "CLOSED"]
        wins = [d for d in closed if d.get("Hasil", "") == "WIN"]
        losses = [d for d in closed if d.get("Hasil", "") == "LOSS"]
        open_signals = [d for d in data if d.get("Status", "") == "OPEN"]

        win_rate = (len(wins) / len(closed) * 100) if closed else 0

        # Hitung profit rata-rata
        profit_list = []
        for d in closed:
            try:
                pl = float(str(d.get("Profit/Loss %", "0")).replace("%", ""))
                profit_list.append(pl)
            except:
                pass

        avg_profit = sum(profit_list) / len(profit_list) if profit_list else 0
        max_loss = min(profit_list) if profit_list else 0

        return {
            "total_sinyal": total,
            "sinyal_open": len(open_signals),
            "sinyal_closed": len(closed),
            "wins": len(wins),
            "losses": len(losses),
            "win_rate": round(win_rate, 1),
            "avg_profit": round(avg_profit, 2),
            "max_drawdown": round(max_loss, 2)
        }

    except Exception as e:
        print(f"⚠️ Gagal hitung performa: {e}")
        return None

def update_sinyal_closed(nama_aset):
    try:
        gc = get_sheets_client()
        if not gc:
            return

        sheet_id = os.getenv("GOOGLE_SHEETS_ID")
        spreadsheet = gc.open_by_key(sheet_id)
        ws = spreadsheet.worksheet("Performance")
        data = ws.get_all_records()

        for i, row in enumerate(data, 2):
            if row["Aset"] == nama_aset.upper() and row["Status"] == "OPEN":
                try:
                    # Ambil harga terkini
                    if "." in nama_aset:
                        harga_skrg = get_stock_price(nama_aset)["harga"]
                    else:
                        data_harga = get_crypto_price(nama_aset)
                        aset_key = list(data_harga.keys())[0]
                        harga_skrg = data_harga[aset_key]["idr"]

                    entry = float(row["Harga Entry"])
                    sl = float(row["Stop Loss"])
                    tp = float(row["Take Profit"])
                    sinyal = row["Sinyal"]

                    # Tentukan WIN/LOSS
                    hasil = "-"
                    pl_pct = 0

                    if sinyal == "BELI":
                        if harga_skrg >= tp:
                            hasil = "WIN"
                            pl_pct = round((tp - entry) / entry * 100, 2)
                        elif harga_skrg <= sl:
                            hasil = "LOSS"
                            pl_pct = round((sl - entry) / entry * 100, 2)

                    elif sinyal == "JUAL":
                        if harga_skrg <= tp:
                            hasil = "WIN"
                            pl_pct = round((entry - tp) / entry * 100, 2)
                        elif harga_skrg >= sl:
                            hasil = "LOSS"
                            pl_pct = round((entry - sl) / entry * 100, 2)

                    if hasil != "-":
                        ws.update_cell(i, 8, harga_skrg)
                        ws.update_cell(i, 9, hasil)
                        ws.update_cell(i, 10, f"{pl_pct}%")
                        ws.update_cell(i, 11, "CLOSED")
                        print(f"✅ Sinyal {nama_aset} diupdate: {hasil} ({pl_pct}%)")

                except Exception as e:
                    print(f"⚠️ Error update sinyal: {e}")

    except Exception as e:
        print(f"⚠️ Gagal update sinyal: {e}")

# ==============================
# FUNGSI SIMPAN KE EXCEL
# ==============================
def simpan_ke_excel(jenis, nama_aset, harga_data, sinyal, analisis):
    nama_file = "histori_analisis.xlsx"
    
    # Cek apakah file sudah ada
    if os.path.exists(nama_file):
        wb = openpyxl.load_workbook(nama_file)
        ws = wb.active
    else:
        # Buat file baru dengan header
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Histori Analisis"
        ws.append([
            "Tanggal & Waktu",
            "Jenis",
            "Nama Aset",
            "Harga (IDR/USD)",
            "Perubahan 24 Jam",
            "Sinyal AI",
            "Analisis Lengkap"
        ])
        # Style header
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
    
    # Ambil data harga
    try:
        if jenis == "Crypto":
            aset_key = list(harga_data.keys())[0]
            harga = f"IDR {harga_data[aset_key].get('idr', '-'):,} | USD {harga_data[aset_key].get('usd', '-'):,}"
            perubahan = f"{harga_data[aset_key].get('idr_24h_change', 0):.2f}%"
        else:
            harga = f"{harga_data.get('mata_uang', '')} {harga_data.get('harga', '-')}"
            perubahan = "-"
    except:
        harga = "-"
        perubahan = "-"
    
    # Tentukan sinyal dari teks analisis
    sinyal_detected = "HOLD"
    analisis_str = analisis if isinstance(analisis, str) else str(analisis)
    if "BELI" in analisis_str.upper():
        sinyal_detected = "BELI"
    elif "JUAL" in analisis_str.upper():
        sinyal_detected = "JUAL"
    
    # Tambah baris baru
    ws.append([
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        jenis,
        nama_aset.upper(),
        harga,
        perubahan,
        sinyal_detected,
        analisis_str[:500]
    ])
    
    # Auto-width kolom
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    wb.save(nama_file)
    print(f"✅ Analisis tersimpan ke {nama_file}!")

# ==============================
# FUNGSI ALERT HARGA
# ==============================
alert_list = []  # Simpan daftar alert

# ==============================
# FUNGSI SIMPAN & LOAD ALERT
# ==============================
ALERT_FILE = "alerts.json"

def simpan_alert():
    with open(ALERT_FILE, "w") as f:
        json.dump(alert_list, f)

def load_alert():
    global alert_list
    if os.path.exists(ALERT_FILE):
        with open(ALERT_FILE, "r") as f:
            alert_list = json.load(f)
        if alert_list:
            print(f"✅ {len(alert_list)} alert dimuat dari file!")

def tambah_alert(jenis, nama_aset, harga_target, kondisi):
    alert_list.append({
        "jenis": jenis,
        "nama_aset": nama_aset,
        "harga_target": harga_target,
        "kondisi": kondisi
    })
    simpan_alert()
    print(f"✅ Alert ditambahkan: {nama_aset} {kondisi} {harga_target}")

def cek_alert():
    for alert in alert_list[:]:
        try:
            if alert["jenis"] == "crypto":
                data = get_crypto_price(alert["nama_aset"])
                aset_key = list(data.keys())[0]
                harga_sekarang = data[aset_key]["usd"]
            else:
                data = get_stock_price(alert["nama_aset"])
                harga_sekarang = data["harga"]

            terpicu = False
            if alert["kondisi"] == "naik" and harga_sekarang >= alert["harga_target"]:
                terpicu = True
            elif alert["kondisi"] == "turun" and harga_sekarang <= alert["harga_target"]:
                terpicu = True

            if terpicu:
                pesan = f"{alert['nama_aset']} sudah {alert['kondisi']} ke {harga_sekarang}!"
                print(f"\n🔔 ALERT: {pesan}")
                notification.notify(
                    title="🚨 CRYPTO & SAHAM ALERT!",
                    message=pesan,
                    timeout=10
                )
                alert_list.remove(alert)
                simpan_alert()

        except:
            pass

def jalankan_monitor():
    schedule.every(1).minutes.do(cek_alert)
    while True:
        schedule.run_pending()
        time.sleep(1)

# ==============================
# MENU UTAMA
# ==============================
def main():
    print("=" * 60)
    print("   CRYPTO & SAHAM AGENT v2.0 — 10-Point Analysis System")
    print("   Powered by OpenRouter AI + Deterministic Scoring")
    print("=" * 60)

    load_alert()
    modal = DEFAULT_MODAL
    
    # Jalankan monitor kalau ada alert tersimpan
    if alert_list:
        t = threading.Thread(target=jalankan_monitor, daemon=True)
        t.start()
        print("🔄 Monitor harga berjalan di background!")
    
    while True:
        print("\nPilih menu:")
        print("1. 📈 Analisis Crypto (10-Point System)")
        print("2. 📉 Analisis Saham (10-Point System)")
        print("3. 🔔 Set Alert Harga")
        print("4. 📋 Lihat Alert Aktif")
        print("5. 🗑️  Hapus Alert")
        print("6. 📊 Lihat Performa")
        print("7. 💰 Set Modal")
        print("8. 🚪 Keluar")

        pilihan = input("\nMasukkan pilihan (1-8): ")

        if pilihan == "1":
            print("\nContoh: bitcoin, ethereum, solana, dogecoin")
            crypto = input("Masukkan nama crypto: ").lower()
            print(f"\n⏳ Mengambil data harga {crypto}...")
            data_harga = get_crypto_price(crypto)
            
            if not data_harga or crypto not in data_harga:
                print("❌ Crypto tidak ditemukan!")
                continue
            
            print(f"✅ Data harga berhasil diambil!")
            print(f"⏳ Mengambil berita terkini {crypto}...")
            berita = get_crypto_news(crypto)
            
            if berita:
                print(f"✅ {len(berita)} berita ditemukan!")
                print("\n📰 Berita Terkini:")
                for i, b in enumerate(berita, 1):
                    print(f"  {i}. {b['judul']}")
            else:
                print("⚠️ Tidak ada berita terkini.")

            print(f"\n⏳ Menghitung indikator teknikal {crypto} (enhanced)...")
            indikator = get_crypto_indicators(crypto)
            if indikator:
                print(f"✅ Indikator berhasil dihitung!")
                print(f"   📈 RSI: {indikator['rsi']} → {indikator['rsi_status']}")
                print(f"   📉 MACD: {indikator['macd_status']}")
                print(f"   📊 Trend: {indikator['trend_status']} (MA50: {indikator.get('ma50', 'N/A')} | MA200: {indikator.get('ma200', 'N/A')})")
                print(f"   📦 Volume: {indikator['volume_status']} (rasio: {indikator['volume_ratio']}x)")
                print(f"   🏪 Market: {indikator['market_condition']}")
            else:
                print("⚠️ Indikator tidak tersedia.")

            print(f"\n🤖 AI sedang menganalisis {crypto} (10-Point System)...\n")
            hasil, analysis_data = analisis_ai_v2("crypto", "Crypto", data_harga, berita, indikator, modal)
            print("-" * 60)
            print(hasil)
            print("-" * 60)
            
            # Simpan ke Excel & Sheets
            simpan_ke_excel("Crypto", crypto, data_harga, "", hasil)
            simpan_ke_sheets("Crypto", crypto, data_harga, hasil, indikator)
            
            # Catat ke performance tracker
            if analysis_data.get("sinyal") != "HOLD":
                catat_sinyal(
                    "Crypto", crypto,
                    analysis_data["entry"],
                    analysis_data["sinyal"],
                    analysis_data["sl"],
                    analysis_data["tp"],
                    analysis_data.get("skor_detail")
                )
            
        elif pilihan == "2":
            print("\nContoh saham Indonesia: BBCA.JK, GOTO.JK, TLKM.JK, BBRI.JK")
            print("Contoh saham US: AAPL, TSLA, NVDA, GOOGL")
            saham = input("Masukkan kode saham: ").upper()
            print(f"\n⏳ Mengambil data harga {saham}...")
            
            try:
                data_harga = get_stock_price(saham)
                print(f"✅ Data harga berhasil diambil!")
                print(f"⏳ Mengambil berita terkini {saham}...")
                berita = get_stock_news(saham)
                
                if berita:
                    print(f"✅ {len(berita)} berita ditemukan!")
                    print("\n📰 Berita Terkini:")
                    for i, b in enumerate(berita, 1):
                        print(f"  {i}. {b['judul']} ({b['sumber']})")
                else:
                    print("⚠️ Tidak ada berita terkini.")

                print(f"\n⏳ Menghitung indikator teknikal {saham} (enhanced)...")
                indikator = get_stock_indicators(saham)
                if indikator:
                    print(f"✅ Indikator berhasil dihitung!")
                    print(f"   📈 RSI: {indikator['rsi']} → {indikator['rsi_status']}")
                    print(f"   📉 MACD: {indikator['macd_status']}")
                    print(f"   📊 Trend: {indikator['trend_status']} (MA50: {indikator.get('ma50', 'N/A')} | MA200: {indikator.get('ma200', 'N/A')})")
                    print(f"   📦 Volume: {indikator['volume_status']} (rasio: {indikator['volume_ratio']}x)")
                    print(f"   🏪 Market: {indikator['market_condition']}")
                else:
                    print("⚠️ Indikator tidak tersedia.")

                print(f"\n🤖 AI sedang menganalisis {saham} (10-Point System)...\n")
                hasil, analysis_data = analisis_ai_v2(saham, "Saham", data_harga, berita, indikator, modal)
                print("-" * 60)
                print(hasil)
                print("-" * 60)
                
                # Simpan
                simpan_ke_excel("Saham", saham, data_harga, "", hasil)
                simpan_ke_sheets("Saham", saham, data_harga, hasil, indikator)
                
                # Catat ke performance tracker
                if analysis_data.get("sinyal") != "HOLD":
                    catat_sinyal(
                        "Saham", saham,
                        analysis_data["entry"],
                        analysis_data["sinyal"],
                        analysis_data["sl"],
                        analysis_data["tp"],
                        analysis_data.get("skor_detail")
                    )
                    
            except Exception as e:
                if "403" in str(e) or "PermissionDenied" in str(type(e).__name__):
                    print(f"❌ API error - coba lagi atau gunakan VPN!")
                else:
                    print(f"❌ Kode saham tidak ditemukan! Pastikan format benar.")
                    print(f"Detail: {e}")
                
        elif pilihan == "3":
            print("\nPilih jenis alert:")
            print("1. Crypto")
            print("2. Saham")
            jenis_alert = input("Pilihan (1/2): ")
            
            if jenis_alert == "1":
                nama = input("Nama crypto (contoh: bitcoin): ").lower()
                jenis = "crypto"
            else:
                nama = input("Kode saham (contoh: BBCA.JK): ").upper()
                jenis = "saham"
            
            kondisi = input("Alert ketika harga (naik/turun): ").lower()
            harga_target = float(input("Harga target (USD untuk crypto, IDR untuk saham): "))
            tambah_alert(jenis, nama, harga_target, kondisi)
            
            # Jalankan monitor di background
            if len(alert_list) == 1:
                t = threading.Thread(target=jalankan_monitor, daemon=True)
                t.start()
                print("🔄 Monitor harga berjalan di background!")

        elif pilihan == "4":
            if not alert_list:
                print("\n⚠️ Tidak ada alert aktif!")
            else:
                print("\n📋 Alert Aktif:")
                for i, a in enumerate(alert_list, 1):
                    print(f"  {i}. {a['nama_aset']} → {a['kondisi']} {a['harga_target']}")

        elif pilihan == "5":
            if not alert_list:
                print("\n⚠️ Tidak ada alert yang bisa dihapus!")
            else:
                print("\n📋 Pilih alert yang ingin dihapus:")
                for i, a in enumerate(alert_list, 1):
                    print(f"  {i}. {a['nama_aset']} → {a['kondisi']} {a['harga_target']}")
                try:
                    nomor = int(input("Masukkan nomor alert: ")) - 1
                    dihapus = alert_list.pop(nomor)
                    simpan_alert()
                    print(f"✅ Alert {dihapus['nama_aset']} berhasil dihapus!")
                except:
                    print("❌ Nomor tidak valid!")

        elif pilihan == "6":
            print("\n⏳ Menghitung performa bot...")
            performa = hitung_performa()
            if not performa:
                print("⚠️ Belum ada data performa!")
            else:
                print("\n" + "=" * 60)
                print("   📊 PERFORMA BOT")
                print("=" * 60)
                print(f"   📈 Total Sinyal : {performa['total_sinyal']}")
                print(f"   🟢 Open         : {performa['sinyal_open']}")
                print(f"   🔵 Closed       : {performa['sinyal_closed']}")
                print(f"   ✅ WIN          : {performa['wins']}")
                print(f"   ❌ LOSS         : {performa['losses']}")
                print(f"   🎯 Win Rate     : {performa['win_rate']}%")
                print(f"   💰 Avg Profit   : {performa['avg_profit']}%")
                print(f"   📉 Max Drawdown : {performa['max_drawdown']}%")
                print("=" * 60)
                
                # Opsi update sinyal
                update = input("\nUpdate sinyal tertentu? (y/n): ").lower()
                if update == "y":
                    nama = input("Masukkan nama aset (contoh: bitcoin / BBCA.JK): ")
                    print(f"⏳ Mengupdate sinyal {nama}...")
                    update_sinyal_closed(nama)

        elif pilihan == "7":
            print(f"\n💰 Modal saat ini: Rp {modal:,.0f}")
            try:
                modal_baru = float(input("Masukkan modal baru (dalam Rupiah): "))
                modal = modal_baru
                print(f"✅ Modal diubah menjadi Rp {modal:,.0f}")
            except:
                print("❌ Input tidak valid!")

        elif pilihan == "8":
            print("\nSampai jumpa! 👋")
            break
        else:
            print("❌ Pilihan tidak valid!")

if __name__ == "__main__":
    main()